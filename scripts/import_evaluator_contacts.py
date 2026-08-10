"""Import evaluator full names and phone numbers into the master directory.

The workbook is authoritative by nickname. Preview is the default; pass
``--confirm`` to write. No contact values are printed to the terminal.
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.db import SessionLocal
from app.models import EvaluatorDirectory, UserAccount
from app.utils import audit


def clean(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def load_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = {clean(cell.value).casefold(): index for index, cell in enumerate(sheet[1])}
        required = {"nickname", "role", "full name", "phone number"}
        missing = sorted(required - set(headers))
        if missing:
            raise ValueError(f"Missing workbook columns: {', '.join(missing)}")
        rows: list[dict[str, str]] = []
        seen: set[str] = set()
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            nickname = clean(row[headers["nickname"]])
            if not nickname:
                continue
            key = nickname.casefold()
            if key in seen:
                raise ValueError(f"Duplicate nickname at row {row_number}: {nickname}")
            seen.add(key)
            role = clean(row[headers["role"]]).casefold()
            if role not in {"overall", "dossard"}:
                raise ValueError(f"Invalid role at row {row_number}: {role or 'blank'}")
            full_name = clean(row[headers["full name"]])
            phone_number = clean(row[headers["phone number"]])
            if not full_name or not phone_number:
                raise ValueError(f"Missing full name or phone number at row {row_number}: {nickname}")
            rows.append(
                {
                    "nickname": nickname,
                    "role": role,
                    "full_name": full_name,
                    "phone_number": phone_number,
                }
            )
        if not rows:
            raise ValueError("The workbook contains no evaluator rows.")
        return rows
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--confirm", action="store_true")
    args = parser.parse_args()
    rows = load_rows(args.workbook)
    with SessionLocal() as db:
        directory = {
            item.name.casefold(): item
            for item in db.scalars(select(EvaluatorDirectory))
        }
        missing = [row["nickname"] for row in rows if row["nickname"].casefold() not in directory]
        if missing:
            raise SystemExit(
                f"Refusing import: {len(missing)} workbook nicknames are absent from the database directory."
            )
        changed = 0
        for row in rows:
            item = directory[row["nickname"].casefold()]
            values = (row["role"], row["full_name"], row["phone_number"])
            before = (item.default_role, item.full_name or "", item.phone_number or "")
            if values != before:
                changed += 1
            item.default_role, item.full_name, item.phone_number = values
            account = db.scalar(select(UserAccount).where(UserAccount.directory_id == item.id))
            if account:
                account.evaluator_role = "dossard" if account.is_owner else row["role"]
                account.version += 1
        print(f"Validated {len(rows)} evaluator contacts; {changed} directory records would change.")
        if not args.confirm:
            print("Preview only. Re-run with --confirm to write changes.")
            db.rollback()
            return
        audit(
            db,
            journey_id=None,
            actor_type="maintenance",
            actor_name="Evaluator contact import",
            action="evaluator_directory.contacts_imported",
            entity_type="evaluator_directory",
            after={"validated": len(rows), "changed": changed, "source": args.workbook.name},
        )
        db.commit()
        print(f"Imported {len(rows)} evaluator contacts successfully.")


if __name__ == "__main__":
    main()
