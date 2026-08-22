from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from decimal import Decimal
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import (ActivityState, Assignment, AssignmentRound, EvaluationSubmission,
                     Evaluator, EvaluatorDirectory, GeneralAssessment, Journey, Recruit,
                     SubmissionVersion)
from .rubric import ACTIVITY_ORDER, RUBRICS
from .scoring import sport_score
from .services import create_journey, process_photo, result_snapshot
from .object_storage import write_recruit_photo
from .utils import audit, dumps

SHEETS = {"sport": "Sport Evaluations", "escape_room": "Escape Room Evaluations",
          "negotiation": "Negotiation Evaluations", "skills": "Skills Evaluations",
          "simulation": "Simulation Evaluations"}
RANK_COLUMNS = {"escape_room": (1, 2), "sport": (3, 4), "negotiation": (5, 6),
                "skills": (7, 8), "simulation": (9, 10)}
BEIRUT = ZoneInfo("Asia/Beirut")
LEGACY_WEIGHTED_CRITERIA = {
    "escape_room": (
        ("active_participation", Decimal("0.15")),
        ("initiative_task_division", Decimal("0.15")),
        ("communication_active_listening", Decimal("0.15")),
        ("self_control_composure", Decimal("0.10")),
        ("time_management", Decimal("0.10")),
        ("setback_recovery", Decimal("0.10")),
        ("logical_problem_solving", Decimal("0.15")),
        ("common_sense_prioritization", Decimal("0.10")),
    ),
    "negotiation": (
        ("initiative_persistence", Decimal("0.15")),
        ("team_support", Decimal("0.10")),
        ("empathy_listening", Decimal("0.10")),
        ("non_aggressive_communication", Decimal("0.15")),
        ("adaptable_behavior", Decimal("0.15")),
        ("self_preservation", Decimal("0.10")),
        ("reasoning_persuasion", Decimal("0.15")),
        ("understanding_boundaries", Decimal("0.10")),
    ),
    "simulation": (
        ("commitment_effort", Decimal("0.15")),
        ("team_cohesion", Decimal("0.15")),
        ("composure_distraction", Decimal("0.15")),
        ("patient_empathy", Decimal("0.10")),
        ("bystander_management", Decimal("0.10")),
        ("recovery_adjustment", Decimal("0.10")),
        ("retention_skills", Decimal("0.10")),
        ("role_assignment", Decimal("0.15")),
    ),
    "skills": (
        ("posture_lifting", Decimal("0.15")),
        ("physical_capability", Decimal("0.15")),
        ("team_synchronization", Decimal("0.10")),
        ("communication_encouragement", Decimal("0.10")),
        ("grasping_concepts", Decimal("0.10")),
        ("correct_placement", Decimal("0.10")),
        ("tension_stability", Decimal("0.10")),
        ("efficiency_speed", Decimal("0.10")),
        ("composure_precision", Decimal("0.10")),
    ),
}


def normalize_name(value: object) -> str:
    return " ".join(str(value or "").split()).casefold()


def clean_name(value: object) -> str:
    return " ".join(str(value or "").split())


def decimal(value: object) -> Decimal:
    return Decimal(str(value))


def legacy_value(value: object) -> object:
    return value.isoformat() if isinstance(value, (datetime, time, date)) else value


def source_timestamp(value: object, event_date: date) -> datetime:
    # Two sheets interpreted 02/08 as February 8. The source filename and the
    # event establish August 2, so preserve the recorded clock time only.
    source_time = value.time() if isinstance(value, datetime) else time(12)
    return datetime.combine(event_date, source_time, tzinfo=BEIRUT).astimezone(timezone.utc)


@dataclass
class EvaluationRow:
    source_row: int
    timestamp: datetime
    recruit: str
    evaluator: str
    responses: dict[str, Decimal]
    raw: dict[str, object]
    score: Decimal
    comments: str


class LegacyWorkbookImporter:
    def __init__(self, workbook_path: str | Path, journey_name: str, event_date: date):
        self.path = Path(workbook_path)
        self.journey_name = clean_name(journey_name)
        self.event_date = event_date
        self.wb = load_workbook(self.path, data_only=True, read_only=False)
        self.duplicate_evaluations: list[dict] = []
        self.over_capacity_assignments: list[dict] = []
        self.duplicate_general: list[str] = []
        self.duplicate_attendance: list[str] = []
        self._recruit_cache: list[dict] | None = None

    def close(self) -> None:
        self.wb.close()

    @staticmethod
    def _rows(sheet, name_col: int):
        for number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if row[name_col] is not None and clean_name(row[name_col]):
                yield number, row

    def recruits(self) -> list[dict]:
        if self._recruit_cache is not None:
            return self._recruit_cache
        sheet = self.wb["Recruits"]
        images = {image.anchor._from.row + 1: image._data() for image in sheet._images}
        result, seen = [], set()
        for number, row in self._rows(sheet, 12):
            name = clean_name(row[12]) or clean_name(f"{row[0] or ''} {row[1] or ''}")
            key = normalize_name(name)
            if key in seen:
                raise ValueError(f"Duplicate recruit name in Recruits: {name}")
            seen.add(key)
            result.append({"row": number, "name": name, "photo": images.get(number)})
        if not result:
            raise ValueError("No recruits found in Recruits.")
        self._recruit_cache = result
        return result

    def evaluations(self) -> dict[str, list[EvaluationRow]]:
        result = {}
        for code, sheet_name in SHEETS.items():
            rows = []
            for number, row in self._rows(self.wb[sheet_name], 1):
                if code == "sport":
                    raw = {"push_ups": row[3], "wall_sit": row[5],
                           "beep_test": row[7], "plank": row[9]}
                    score, _normalized, responses = sport_score(raw)
                    source_score, comments = decimal(row[10]), clean_name(row[11])
                    raw["legacy_entered"] = {"wall_sit": legacy_value(row[4]),
                                               "beep_test": legacy_value(row[6]),
                                               "plank": legacy_value(row[8])}
                else:
                    criteria = LEGACY_WEIGHTED_CRITERIA[code]
                    responses = {key: decimal(row[3 + i]) for i, (key, _weight) in enumerate(criteria)}
                    score = sum((responses[key] * weight for key, weight in criteria), Decimal("0"))
                    source_score = decimal(row[3 + len(criteria)])
                    raw, comments = {}, clean_name(row[4 + len(criteria)])
                if abs(score - source_score) > Decimal("0.011"):
                    raise ValueError(f"{sheet_name} row {number}: calculated {score} != source {source_score}.")
                rows.append(EvaluationRow(number, source_timestamp(row[0], self.event_date),
                                          clean_name(row[1]), clean_name(row[2]), responses,
                                          raw, source_score, comments))
            result[code] = rows
        return result

    def attendance(self) -> dict[str, tuple[str, str]]:
        result = {}
        for _number, row in self._rows(self.wb["203 Attendance"], 0):
            name, role = clean_name(row[0]), clean_name(row[1]).lower()
            role = role if role in {"overall", "dossard"} else "dossard"
            key = normalize_name(name)
            if key in result:
                self.duplicate_attendance.append(name)
            result[key] = (name, role)
        return result

    def assignment_pairs(self) -> list[tuple[str, str]]:
        return [(clean_name(row[1]), clean_name(row[3]))
                for _number, row in self._rows(self.wb["Assignments"], 1)]

    def general_assessments(self) -> dict[str, dict]:
        result = {}
        for number, row in self._rows(self.wb["General Evaluations"], 0):
            name, key = clean_name(row[0]), normalize_name(row[0])
            value = {"name": name, "punctuality": decimal(row[1]), "respect": decimal(row[2]),
                     "seriousness": decimal(row[3]), "legacy_total": decimal(row[4]),
                     "comment": clean_name(row[5]), "source_rows": [number]}
            if key in result:
                old = result[key]
                if any(old[field] != value[field] for field in ("punctuality", "respect", "seriousness")):
                    raise ValueError(f"Conflicting general assessments for {name}.")
                old["source_rows"].append(number)
                old["comment"] = value["comment"] or old["comment"]
                old["legacy_total"] = value["legacy_total"]
                self.duplicate_general.append(name)
            else:
                result[key] = value
        return result

    def source_rankings(self) -> dict[str, dict]:
        result = {}
        for _number, row in self._rows(self.wb["Recruit Rankings"], 0):
            entry = {"name": clean_name(row[0]), "activities": {}}
            for code, (score_column, rank_column) in RANK_COLUMNS.items():
                entry["activities"][code] = {"score": float(row[score_column]), "rank": int(row[rank_column])}
            entry.update(legacy_total_5=float(row[11]), legacy_rank=int(row[12]))
            result[normalize_name(row[0])] = entry
        return result

    def analyze(self) -> dict:
        recruits, evaluations, attendance = self.recruits(), self.evaluations(), self.attendance()
        general = self.general_assessments()
        recruit_keys = {normalize_name(item["name"]) for item in recruits}
        names = {normalize_name(row.evaluator): row.evaluator for rows in evaluations.values() for row in rows}
        names.update({key: value[0] for key, value in attendance.items()})
        unknown = sorted({row.recruit for rows in evaluations.values() for row in rows
                          if normalize_name(row.recruit) not in recruit_keys}
                         | {value["name"] for key, value in general.items() if key not in recruit_keys})
        if unknown:
            raise ValueError("Unknown recruits referenced: " + ", ".join(unknown))
        source_pairs = {(normalize_name(e), normalize_name(r)) for e, r in self.assignment_pairs()}
        skill_pairs = {(normalize_name(row.evaluator), normalize_name(row.recruit)) for row in evaluations["skills"]}
        if source_pairs != skill_pairs:
            raise ValueError("Assignments does not match the Skills evaluator/recruit pairing.")
        return {"journeyName": self.journey_name, "eventDate": self.event_date.isoformat(),
                "recruits": len(recruits), "photos": sum(bool(item["photo"]) for item in recruits),
                "evaluators": len(names), "presentEvaluators": len(attendance),
                "historicalEvaluatorsCurrentlyAbsent": sorted(name for key, name in names.items() if key not in attendance),
                "evaluations": {code: len(rows) for code, rows in evaluations.items()},
                "generalAssessments": len(general),
                "duplicates": {"attendance": sorted(set(self.duplicate_attendance)),
                               "general": sorted(set(self.duplicate_general))}}

    def import_into(self, db: Session, actor_name: str = "Historical Excel import") -> tuple[Journey, dict]:
        if db.scalar(select(Journey).where(Journey.name == self.journey_name,
                                           Journey.event_date == self.event_date)):
            raise ValueError(f"Journee '{self.journey_name}' on {self.event_date} already exists.")
        report = self.analyze()
        recruit_source, evaluation_rows = self.recruits(), self.evaluations()
        attendance, general, rankings = self.attendance(), self.general_assessments(), self.source_rankings()
        journey = create_journey(db, self.journey_name, self.event_date, 1, actor_name)
        journey.status, journey.current_activity = "completed", None
        db.flush()

        recruits = {}
        for source in recruit_source:
            recruit = Recruit(journey_id=journey.id, name=source["name"], present=True, active=True)
            if source["photo"]:
                photo, photo_type = process_photo(source["photo"])
                db.add(recruit); db.flush()
                write_recruit_photo(recruit, photo, photo_type)
            else:
                db.add(recruit); db.flush()
            recruits[normalize_name(recruit.name)] = recruit

        names = {normalize_name(row.evaluator): row.evaluator for rows in evaluation_rows.values() for row in rows}
        names.update({key: value[0] for key, value in attendance.items()})
        directory = {normalize_name(item.name): item for item in db.scalars(select(EvaluatorDirectory))}
        evaluators = {}
        for key, name in names.items():
            role = attendance[key][1] if key in attendance else "dossard"
            directory_item = directory.get(key)
            if directory_item is None:
                directory_item = EvaluatorDirectory(name=name, default_role=role)
                db.add(directory_item); db.flush(); directory[key] = directory_item
            evaluator = Evaluator(journey_id=journey.id, directory_id=directory_item.id, name=name,
                                  role=role, present=key in attendance, active=True)
            db.add(evaluator); db.flush(); evaluators[key] = evaluator

        last_pairs, rounds, activity_times = set(), {}, defaultdict(list)
        assignment_sheet_pairs = self.assignment_pairs()
        for code in ACTIVITY_ORDER:
            rows = evaluation_rows[code]
            raw_pairs = assignment_sheet_pairs if code in {"skills", "simulation"} else [(r.evaluator, r.recruit) for r in rows]
            pairs, seen = [], set()
            for evaluator_name, recruit_name in raw_pairs:
                key = (normalize_name(evaluator_name), normalize_name(recruit_name))
                if key not in seen: seen.add(key); pairs.append(key)
            round_record = AssignmentRound(journey_id=journey.id, activity_code=code, version=1,
                status="published", seed="historical-workbook",
                warnings_json=dumps(["Historical import: room membership was not available in the source workbook."]),
                reused_from_id=rounds["skills"].id if code == "simulation" else None,
                created_by=actor_name,
                published_at=max((row.timestamp for row in rows), default=datetime.now(timezone.utc)))
            db.add(round_record); db.flush(); rounds[code] = round_record
            slots, assignments = defaultdict(int), {}
            for pair in pairs:
                evaluator_key, recruit_key = pair; slots[recruit_key] += 1
                if slots[recruit_key] > 2:
                    # Completed legacy workbooks are imported faithfully even when
                    # they predate the current two-evaluator rule. New previews and
                    # manual edits continue to enforce the live limit.
                    self.over_capacity_assignments.append({
                        "activity": code,
                        "recruit": recruits[recruit_key].name,
                        "evaluator": evaluators[evaluator_key].name,
                        "slot": slots[recruit_key],
                    })
                repeated = pair in last_pairs and code != "simulation"
                assignment = Assignment(round_id=round_record.id, evaluator_id=evaluators[evaluator_key].id,
                    recruit_id=recruits[recruit_key].id, slot=slots[recruit_key], repeated_pair=repeated,
                    repeat_reason="Historical pairing preserved from source workbook." if repeated else None)
                db.add(assignment); db.flush(); assignments[pair] = assignment

            grouped = defaultdict(list)
            for row in rows:
                grouped[(normalize_name(row.evaluator), normalize_name(row.recruit))].append(row)
                activity_times[code].append(row.timestamp)
            for pair, versions in grouped.items():
                versions.sort(key=lambda item: (item.timestamp, item.source_row)); latest = versions[-1]
                imported_score = sum((item.score for item in versions), Decimal("0")) / Decimal(len(versions))
                if len(versions) > 1:
                    self.duplicate_evaluations.append({"activity": code, "evaluator": latest.evaluator,
                        "recruit": latest.recruit, "sourceRows": [item.source_row for item in versions]})
                assignment = assignments[pair]
                submission = EvaluationSubmission(assignment_id=assignment.id, journey_id=journey.id,
                    activity_code=code, evaluator_id=assignment.evaluator_id, recruit_id=assignment.recruit_id,
                    status="locked", responses_json=dumps(latest.responses), raw_payload_json=dumps(latest.raw),
                    comments=latest.comments, score=imported_score, version=len(versions),
                    submitted_at=latest.timestamp, updated_at=latest.timestamp)
                db.add(submission); db.flush()
                for version_number, item in enumerate(versions, 1):
                    db.add(SubmissionVersion(submission_id=submission.id, version=version_number,
                        payload_json=dumps({"responses": item.responses, "raw": item.raw, "comments": item.comments,
                                           "status": "submitted", "legacySourceRow": item.source_row}),
                        score=item.score, actor_type="evaluator", actor_name=item.evaluator,
                        reason=f"Imported from {self.path.name}, row {item.source_row}", created_at=item.timestamp))
            last_pairs.update(pairs)
            state = db.scalar(select(ActivityState).where(ActivityState.journey_id == journey.id,
                                                           ActivityState.code == code))
            state.status, state.assignment_round_id = "closed", round_record.id
            state.opened_at = min(activity_times[code]) if activity_times[code] else round_record.published_at
            state.closed_at = max(activity_times[code]) if activity_times[code] else round_record.published_at

        for key, item in general.items():
            recruit = recruits[key]
            db.add(GeneralAssessment(recruit_id=recruit.id, punctuality=item["punctuality"],
                respect=item["respect"], seriousness=item["seriousness"], comment=item["comment"]))
            audit(db, journey_id=journey.id, actor_type="admin", actor_name=actor_name,
                  action="recruit.profile_imported", entity_type="recruit", entity_id=recruit.id,
                  after={"sourceRows": item["source_rows"], "legacyTotal": item["legacy_total"]},
                  reason="Imported historical assessment; current formula is recalculated from its components.")

        db.flush(); results = result_snapshot(db, journey)
        result_by_name = {normalize_name(row["name"]): row for row in results["rows"]}
        reconciliation, rank_differences = [], []
        for key, source in rankings.items():
            imported = result_by_name.get(key)
            if not imported: raise ValueError(f"Ranking recruit not found: {source['name']}")
            for code, expected in source["activities"].items():
                actual = imported["activities"][code]
                if abs(Decimal(str(actual["score"])) - Decimal(str(expected["score"]))) > Decimal("0.011"):
                    raise ValueError(f"Score reconciliation failed for {source['name']} / {code}.")
                if actual["rank"] != expected["rank"]:
                    rank_differences.append({"recruit": source["name"], "activity": code,
                                             "legacyRank": expected["rank"], "currentRank": actual["rank"]})
            reconciliation.append({"recruit": source["name"], "activityRanksMatched": True,
                "legacyTotal5": source["legacy_total_5"], "legacyRank": source["legacy_rank"],
                "newOverall20": imported["overallScore"], "newOverallRank": imported["overallRank"]})
        report.update(duplicateEvaluations=self.duplicate_evaluations,
                      historicalOverCapacityAssignments=self.over_capacity_assignments,
                      activityScoresReconciled=len(reconciliation), legacyRankDifferences=rank_differences,
                      journeyId=journey.id)
        audit(db, journey_id=journey.id, actor_type="admin", actor_name=actor_name,
              action="historical_import.completed", entity_type="journey", entity_id=journey.id,
              after={**report, "sourceRankReconciliation": reconciliation},
              reason=f"Imported from {self.path.name}. Legacy overall /5 ranks are audited; overall /20 is recalculated.")
        return journey, report
