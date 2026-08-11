from __future__ import annotations

import io
import re
from copy import deepcopy
from collections import Counter, defaultdict
from datetime import datetime, timezone
from decimal import Decimal
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.chart import RadarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import (
    ActivityState,
    AdminEvaluation,
    Assignment,
    AssignmentRound,
    AuditEvent,
    EvaluationSubmission,
    Evaluator,
    GeneralAssessment,
    Journey,
    MandatoryRoomEvaluator,
    Recruit,
    RoomPlanEvaluator,
    RoomPlanRecruit,
)
from .rubric import ACTIVITY_ORDER, DIMENSION_NAMES, DIMENSION_ORDER, RUBRICS
from .scoring import competition_ranks
from .services import latest_room_plan, result_snapshot
from .utils import loads


BEIRUT = ZoneInfo("Asia/Beirut")
NAVY = "223449"
RED = "C8102E"
RED_LIGHT = "FCE9ED"
BLUE_LIGHT = "EAF1F8"
GREEN = "16834B"
GREEN_LIGHT = "E8F5ED"
YELLOW = "E3AD22"
YELLOW_LIGHT = "FFF5D9"
GRAY = "667085"
GRAY_LIGHT = "F3F5F7"
LINE = "D6DBE1"
WHITE = "FFFFFF"
THIN_LINE = Side(style="thin", color=LINE)


def _label(value: object) -> str:
    return str(value or "").replace("_", " ").replace(".", " › ").title()


def _local_time(value: datetime | None) -> str:
    if not value:
        return "Not recorded"
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(BEIRUT).strftime("%d %b %Y, %H:%M")


def _new_sheet(workbook: Workbook, title: str, *, landscape: bool = False):
    sheet = workbook.create_sheet(title=title)
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    sheet.freeze_panes = "A4"
    sheet.page_setup.orientation = "landscape" if landscape else "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.oddFooter.center.text = "LRC Journee Recruitment · View-only export"
    sheet.oddFooter.right.text = "Page &P of &N"
    return sheet


def _title(sheet, title: str, subtitle: str, span: int = 10) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = sheet.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    sub = sheet.cell(2, 1, subtitle)
    sub.font = Font(name="Aptos", size=10, color=GRAY)
    sub.alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 24


def _section(sheet, row: int, title: str, span: int = 10) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = sheet.cell(row, 1, title)
    cell.fill = PatternFill("solid", fgColor=RED_LIGHT)
    cell.font = Font(name="Aptos", size=11, bold=True, color=RED)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row].height = 25
    return row + 1


def _write_table(
    sheet,
    row: int,
    headers: list[str],
    rows: list[list],
    *,
    widths: list[float] | None = None,
    number_formats: dict[int, str] | None = None,
    auto_filter: bool = False,
) -> int:
    header_row = row
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(row, column, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    sheet.row_dimensions[row].height = 28
    for values in rows:
        row += 1
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row, column, value)
            cell.font = Font(name="Aptos", size=9, color="20252B")
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(value, str) and len(value) > 32)
            cell.border = Border(bottom=THIN_LINE)
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFBFC")
            if number_formats and column in number_formats and value not in (None, ""):
                cell.number_format = number_formats[column]
        sheet.row_dimensions[row].height = 22
    if auto_filter and rows:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{row}"
    if widths:
        for column, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(column)].width = width
    return row + 2


def _status_fill(cell, status: str) -> None:
    status = status.casefold()
    if status in {"complete", "completed", "present", "submitted", "locked", "green", "active", "open"}:
        cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
        cell.font = Font(name="Aptos", size=9, bold=True, color=GREEN)
    elif status in {"yellow", "incomplete", "draft", "ready", "not started", "not_started"}:
        cell.fill = PatternFill("solid", fgColor=YELLOW_LIGHT)
        cell.font = Font(name="Aptos", size=9, bold=True, color="745300")
    elif status in {"red", "absent", "missing", "archived", "closed"}:
        cell.fill = PatternFill("solid", fgColor=RED_LIGHT)
        cell.font = Font(name="Aptos", size=9, bold=True, color=RED)


def _add_radar(sheet, start_row: int, end_row: int, category_column: int, value_column: int, anchor: str, title: str, maximum: float) -> None:
    chart = RadarChart()
    data = Reference(sheet, min_col=value_column, min_row=start_row - 1, max_row=end_row)
    categories = Reference(sheet, min_col=category_column, min_row=start_row, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = title
    chart.style = 26
    chart.height = 7.2
    chart.width = 10.5
    chart.legend = None
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = maximum
    sheet.add_chart(chart, anchor)


def _safe_profile_title(workbook: Workbook, index: int, name: str) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", f"Profile {index:02d} - {name}")[:31].strip()
    title = base
    suffix = 2
    while title in workbook.sheetnames:
        marker = f" {suffix}"
        title = (base[: 31 - len(marker)] + marker).strip()
        suffix += 1
    return title


def _collect(db: Session, journey: Journey) -> dict:
    recruits = list(db.scalars(select(Recruit).where(Recruit.journey_id == journey.id, Recruit.active.is_(True)).order_by(Recruit.name)))
    evaluators = list(db.scalars(select(Evaluator).where(Evaluator.journey_id == journey.id, Evaluator.active.is_(True)).order_by(Evaluator.name)))
    recruit_by_id = {item.id: item for item in recruits}
    evaluator_by_id = {item.id: item for item in evaluators}
    states = {item.code: item for item in db.scalars(select(ActivityState).where(ActivityState.journey_id == journey.id))}
    round_ids = [item.assignment_round_id for item in states.values() if item.assignment_round_id]
    rounds = {item.id: item for item in db.scalars(select(AssignmentRound).where(AssignmentRound.id.in_(round_ids)))} if round_ids else {}
    assignments = list(db.scalars(select(Assignment).where(Assignment.round_id.in_(round_ids)))) if round_ids else []
    assignment_by_id = {item.id: item for item in assignments}
    assignment_activity = {item.id: rounds[item.round_id].activity_code for item in assignments if item.round_id in rounds}
    submission_rows = list(db.scalars(select(EvaluationSubmission).where(EvaluationSubmission.assignment_id.in_(list(assignment_by_id))))) if assignments else []
    submissions = {item.assignment_id: item for item in submission_rows}
    admin_evaluations = list(db.scalars(select(AdminEvaluation).where(AdminEvaluation.journey_id == journey.id)))
    assessments = {item.recruit_id: item for item in db.scalars(select(GeneralAssessment).where(GeneralAssessment.recruit_id.in_(list(recruit_by_id))))} if recruits else {}
    results = result_snapshot(db, journey)
    result_by_recruit = {item["recruitId"]: item for item in results["rows"]}
    return {
        "recruits": recruits,
        "evaluators": evaluators,
        "recruit_by_id": recruit_by_id,
        "evaluator_by_id": evaluator_by_id,
        "states": states,
        "rounds": rounds,
        "assignments": assignments,
        "assignment_activity": assignment_activity,
        "submissions": submissions,
        "admin_evaluations": admin_evaluations,
        "assessments": assessments,
        "results": results,
        "result_by_recruit": result_by_recruit,
    }


def _dashboard_sheet(workbook: Workbook, journey: Journey, data: dict) -> None:
    sheet = _new_sheet(workbook, "Journee Dashboard", landscape=True)
    _title(sheet, journey.name, f"{journey.event_date:%d %B %Y} · {_label(journey.status)} · Exported {_local_time(datetime.now(timezone.utc))}", 12)
    recruits = data["recruits"]
    evaluators = data["evaluators"]
    metrics = [
        ("Present recruits", sum(item.present for item in recruits), len(recruits)),
        ("Present evaluators", sum(item.present for item in evaluators), len(evaluators)),
        ("Overall evaluators", sum(item.present and item.role == "overall" for item in evaluators), "present"),
        ("Dossard evaluators", sum(item.present and item.role == "dossard" for item in evaluators), "present"),
    ]
    for index, (label, value, context) in enumerate(metrics):
        column = 1 + index * 3
        sheet.merge_cells(start_row=4, start_column=column, end_row=4, end_column=column + 1)
        sheet.merge_cells(start_row=5, start_column=column, end_row=6, end_column=column + 1)
        sheet.cell(4, column, label).fill = PatternFill("solid", fgColor=BLUE_LIGHT)
        sheet.cell(4, column).font = Font(name="Aptos", size=9, bold=True, color=NAVY)
        metric_text = f"{value} of {context} present" if isinstance(context, int) else f"{value} {context}"
        sheet.cell(5, column, metric_text).font = Font(name="Aptos Display", size=15, bold=True, color=RED)
        sheet.cell(5, column).alignment = Alignment(vertical="center")
    row = _section(sheet, 8, "Activity lifecycle and completion", 12)
    activity_rows = []
    for code in ACTIVITY_ORDER:
        state = data["states"].get(code)
        assignments = [item for item in data["assignments"] if data["assignment_activity"].get(item.id) == code]
        submitted = sum(item.id in data["submissions"] and data["submissions"][item.id].status in {"submitted", "locked"} for item in assignments)
        activity_rows.append([RUBRICS[code].name, _label(state.status if state else "not_started"), len(assignments), submitted, len(assignments) - submitted, data["results"]["activityAverages"].get(code, 0)])
    row = _write_table(sheet, row, ["Activity", "Status", "Expected evaluations", "Received", "Missing", "Average /5"], activity_rows, widths=[24, 16, 20, 13, 13, 14], number_formats={6: "0.00"})
    for item_row in range(10, 10 + len(activity_rows)):
        _status_fill(sheet.cell(item_row, 2), str(sheet.cell(item_row, 2).value))
    row = _section(sheet, row, "Dimension and activity averages", 12)
    averages = [[DIMENSION_NAMES[code], data["results"]["dimensionAverages"].get(code, 0), "Dimension /1"] for code in DIMENSION_ORDER]
    averages.extend([[RUBRICS[code].name, data["results"]["activityAverages"].get(code, 0), "Activity /5"] for code in ACTIVITY_ORDER])
    row = _write_table(sheet, row, ["Measure", "Average", "Scale"], averages, widths=[28, 14, 16], number_formats={2: "0.00"})
    row = _section(sheet, row, "Provisional overall ranking", 12)
    ranking_rows = [[item["overallRank"], item["name"], item["overallScore"], item["color"].title(), item["missingCount"]] for item in data["results"]["rows"]]
    _write_table(sheet, row, ["Rank", "Recruit", "Overall /20", "Color", "Missing components"], ranking_rows, widths=[10, 30, 16, 13, 20], number_formats={3: "0.00"}, auto_filter=True)


def _attendance_sheets(workbook: Workbook, journey: Journey, data: dict) -> None:
    recruits = _new_sheet(workbook, "Recruit Attendance", landscape=True)
    _title(recruits, "Recruit Attendance", journey.name, 8)
    rows = [[item.name, "Present" if item.present else "Absent", _local_time(item.arrival_time), item.attendance_comment or "", item.phone_number or "", item.date_of_birth, "Yes" if item.photo_data else "No"] for item in data["recruits"]]
    _write_table(recruits, 4, ["Recruit", "Attendance", "Time of arrival (Beirut)", "Attendance comment", "Phone number", "Date of birth", "Photo"], rows, widths=[30, 14, 25, 30, 18, 16, 10], auto_filter=True)
    for row in range(5, 5 + len(rows)):
        _status_fill(recruits.cell(row, 2), str(recruits.cell(row, 2).value))
        recruits.cell(row, 6).number_format = "dd mmm yyyy"

    evaluators = _new_sheet(workbook, "Evaluator Attendance")
    _title(evaluators, "Evaluator Attendance", journey.name, 6)
    rows = [[item.name, item.role.title(), "Present" if item.present else "Absent"] for item in sorted(data["evaluators"], key=lambda item: (not item.present, item.role != "overall", item.name.casefold()))]
    _write_table(evaluators, 4, ["Evaluator", "Role", "Attendance"], rows, widths=[32, 16, 16], auto_filter=True)
    for row in range(5, 5 + len(rows)):
        _status_fill(evaluators.cell(row, 3), str(evaluators.cell(row, 3).value))


def _rooms_sheet(workbook: Workbook, journey: Journey, data: dict) -> None:
    sheet = _new_sheet(workbook, "Rooms", landscape=True)
    _title(sheet, "Published Room Distribution", journey.name, 8)
    plan = latest_room_plan(data["db"], journey.id, "published")
    if not plan:
        sheet.cell(4, 1, "No published room distribution.")
        return
    recruits_by_room: dict[int, list[str]] = defaultdict(list)
    evaluators_by_room: dict[int, list[str]] = defaultdict(list)
    for member in data["db"].scalars(select(RoomPlanRecruit).where(RoomPlanRecruit.plan_id == plan.id)):
        recruit = data["recruit_by_id"].get(member.recruit_id)
        recruits_by_room[member.room_number].append(recruit.name if recruit else "Unknown")
    for member in data["db"].scalars(select(RoomPlanEvaluator).where(RoomPlanEvaluator.plan_id == plan.id)):
        evaluator = data["evaluator_by_id"].get(member.evaluator_id)
        label = f"{evaluator.name if evaluator else 'Unknown'} ({(evaluator.role if evaluator else '').title()})"
        if member.mandatory:
            label += " · Mandatory"
        evaluators_by_room[member.room_number].append(label)
    row = 4
    room_count = max([journey.room_count, *recruits_by_room.keys(), *evaluators_by_room.keys()])
    for room in range(1, room_count + 1):
        row = _section(sheet, row, f"Room {room}", 8)
        maximum = max(len(recruits_by_room[room]), len(evaluators_by_room[room]), 1)
        room_rows = [[recruits_by_room[room][index] if index < len(recruits_by_room[room]) else "", evaluators_by_room[room][index] if index < len(evaluators_by_room[room]) else ""] for index in range(maximum)]
        row = _write_table(sheet, row, ["Recruits", "Evaluators"], room_rows, widths=[34, 42])


def _assignments_sheet(workbook: Workbook, journey: Journey, data: dict) -> None:
    sheet = _new_sheet(workbook, "Assignments", landscape=True)
    _title(sheet, "Current Published Assignments", journey.name, 8)
    rows = []
    for assignment in sorted(data["assignments"], key=lambda item: (ACTIVITY_ORDER.index(data["assignment_activity"].get(item.id, "sport")), item.room_number or 0, data["recruit_by_id"].get(item.recruit_id).name.casefold() if data["recruit_by_id"].get(item.recruit_id) else "")):
        code = data["assignment_activity"].get(assignment.id)
        recruit = data["recruit_by_id"].get(assignment.recruit_id)
        evaluator = data["evaluator_by_id"].get(assignment.evaluator_id)
        submission = data["submissions"].get(assignment.id)
        rows.append([RUBRICS[code].name if code else "Unknown", assignment.room_number or "—", recruit.name if recruit else "Unknown", evaluator.name if evaluator else "Unknown", evaluator.role.title() if evaluator else "", assignment.slot, _label(submission.status) if submission else "Missing", float(submission.score) if submission else ""])
    _write_table(sheet, 4, ["Activity", "Room", "Recruit", "Evaluator", "Role", "Slot", "Evaluation status", "Score /5"], rows, widths=[20, 9, 28, 28, 13, 9, 18, 13], number_formats={8: "0.00"}, auto_filter=True)


def _evaluation_sheet(workbook: Workbook, journey: Journey, data: dict) -> None:
    sheet = _new_sheet(workbook, "Evaluation Details", landscape=True)
    _title(sheet, "Evaluation Details", f"Human-readable criterion report · {journey.name}", 10)
    rows = []
    for assignment in data["assignments"]:
        submission = data["submissions"].get(assignment.id)
        if not submission:
            continue
        code = data["assignment_activity"].get(assignment.id)
        recruit = data["recruit_by_id"].get(assignment.recruit_id)
        evaluator = data["evaluator_by_id"].get(assignment.evaluator_id)
        responses = loads(submission.responses_json, {})
        raw = loads(submission.raw_payload_json, {})
        for criterion in RUBRICS[code].criteria:
            grade = responses.get(criterion.key)
            result = raw.get(criterion.key, "")
            rows.append([RUBRICS[code].name, recruit.name if recruit else "Unknown", evaluator.name if evaluator else "Unknown", evaluator.role.title() if evaluator else "", _label(submission.status), float(submission.score), criterion.dimension, criterion.name, grade if grade is not None else "", f"{result} {criterion.unit}".strip() if result != "" else "", submission.comments])
    for evaluation in data["admin_evaluations"]:
        recruit = data["recruit_by_id"].get(evaluation.recruit_id)
        responses = loads(evaluation.responses_json, {})
        raw = loads(evaluation.raw_payload_json, {})
        for criterion in RUBRICS[evaluation.activity_code].criteria:
            result = raw.get(criterion.key, "")
            rows.append([RUBRICS[evaluation.activity_code].name, recruit.name if recruit else "Unknown", evaluation.updated_by, "Admin", "Official admin evaluation", float(evaluation.score), criterion.dimension, criterion.name, responses.get(criterion.key, ""), f"{result} {criterion.unit}".strip() if result != "" else "", evaluation.comments])
    _write_table(sheet, 4, ["Activity", "Recruit", "Evaluator", "Role", "Status", "Evaluation /5", "Dimension", "Criterion", "Grade /5", "Sport result", "Comment"], rows, widths=[18, 26, 26, 12, 14, 15, 18, 38, 12, 16, 42], number_formats={6: "0.00", 9: "0.00"}, auto_filter=True)


def _rubric_sheet(workbook: Workbook, journey: Journey) -> None:
    sheet = _new_sheet(workbook, "Rubric Guide", landscape=True)
    _title(sheet, "2026 Evaluation Rubric", f"Read-only scoring reference · {journey.name}", 9)
    rows = []
    for code in ACTIVITY_ORDER:
        for criterion in RUBRICS[code].criteria:
            rows.append([RUBRICS[code].name, criterion.dimension, criterion.name, float(criterion.weight), criterion.explanation, _label(criterion.input_type), float(criterion.target) if criterion.target is not None else "", criterion.unit])
    _write_table(sheet, 4, ["Activity", "Dimension / theme", "Criterion", "Activity weight", "Explanation", "Input", "Target", "Unit"], rows, widths=[18, 20, 36, 16, 60, 14, 12, 14], number_formats={4: "0%"}, auto_filter=True)


def _results_sheets(workbook: Workbook, journey: Journey, data: dict) -> None:
    summary = _new_sheet(workbook, "Results Summary", landscape=True)
    _title(summary, "Overall Results & Rankings", f"{journey.name} · {data['results']['formula']}", 16)
    headers = ["Rank", "Recruit", "Overall /20", "Color", "Missing", *[DIMENSION_NAMES[code] + " /1" for code in DIMENSION_ORDER], "General /1", *[RUBRICS[code].name + " /5" for code in ACTIVITY_ORDER]]
    rows = [[item["overallRank"], item["name"], item["overallScore"], item["color"].title(), item["missingCount"], *[item["dimensions"][code]["score"] for code in DIMENSION_ORDER], item["generalAverage"], *[item["activities"][code]["score"] for code in ACTIVITY_ORDER]] for item in data["results"]["rows"]]
    _write_table(summary, 4, headers, rows, widths=[8, 27, 14, 11, 10, *([15] * len(DIMENSION_ORDER)), 13, *([14] * len(ACTIVITY_ORDER))], number_formats={column: "0.00" for column in range(3, len(headers) + 1) if column not in {4, 5}}, auto_filter=True)
    for row in range(5, 5 + len(rows)):
        _status_fill(summary.cell(row, 4), str(summary.cell(row, 4).value))

    dimensions = _new_sheet(workbook, "Dimension Rankings", landscape=True)
    _title(dimensions, "Dimension Rankings", journey.name, 7)
    row = 4
    for code in DIMENSION_ORDER:
        row = _section(dimensions, row, f"{DIMENSION_NAMES[code]} · Average {data['results']['dimensionAverages'].get(code, 0):.2f} /1", 7)
        ranking = sorted(data["results"]["rows"], key=lambda item: (item["dimensions"][code]["rank"] or 10**9, item["name"]))
        rows = [[item["dimensions"][code]["rank"], item["name"], item["dimensions"][code]["score"], item["dimensions"][code]["availableWeight"], "Complete" if item["dimensions"][code]["complete"] else "Incomplete"] for item in ranking]
        row = _write_table(dimensions, row, ["Rank", "Recruit", "Grade /1", "Coverage", "Status"], rows, widths=[9, 30, 14, 13, 15], number_formats={3: "0.00", 4: "0%"})

    activities = _new_sheet(workbook, "Activity Rankings", landscape=True)
    _title(activities, "Activity Rankings", journey.name, 7)
    row = 4
    for code in ACTIVITY_ORDER:
        row = _section(activities, row, f"{RUBRICS[code].name} · Average {data['results']['activityAverages'].get(code, 0):.2f} /5", 7)
        ranking = sorted(data["results"]["rows"], key=lambda item: (item["activities"][code]["rank"] or 10**9, item["name"]))
        rows = [[item["activities"][code]["rank"], item["name"], item["activities"][code]["score"], item["activities"][code]["submitted"], item["activities"][code]["expected"], "Complete" if item["activities"][code]["complete"] else "Incomplete"] for item in ranking]
        row = _write_table(activities, row, ["Rank", "Recruit", "Grade /5", "Submitted", "Expected", "Status"], rows, widths=[9, 30, 14, 13, 13, 15], number_formats={3: "0.00"})


def _profile_sheet(workbook: Workbook, journey: Journey, data: dict, recruit: Recruit, index: int) -> None:
    title = _safe_profile_title(workbook, index, recruit.name)
    sheet = _new_sheet(workbook, title, landscape=True)
    result = data["result_by_recruit"].get(recruit.id)
    assessment = data["assessments"].get(recruit.id)
    color = (result or {}).get("color", "red")
    _title(sheet, recruit.name, f"Recruit profile · {journey.name}", 12)
    details = [
        ["Attendance", "Present" if recruit.present else "Absent", "Arrival", _local_time(recruit.arrival_time)],
        ["Phone", recruit.phone_number or "Not recorded", "Date of birth", recruit.date_of_birth or "Not recorded"],
        ["Overall /20", (result or {}).get("overallScore", 0), "Overall rank", (result or {}).get("overallRank", "—")],
        ["Color grade", color.title(), "Missing components", (result or {}).get("missingCount", 8)],
    ]
    if recruit.attendance_comment:
        details.append(["Attendance comment", recruit.attendance_comment, "", ""])
    row = _write_table(sheet, 4, ["Profile field", "Value", "Profile field", "Value"], details, widths=[18, 25, 20, 25])
    _status_fill(sheet.cell(8, 2), color)
    if recruit.photo_data:
        try:
            source = io.BytesIO(recruit.photo_data)
            output = io.BytesIO()
            with PillowImage.open(source) as image:
                image.convert("RGB").save(output, format="PNG")
            output.seek(0)
            photo = ExcelImage(output)
            photo.width = 145
            photo.height = 145
            sheet.add_image(photo, "F4")
        except Exception:
            sheet.cell(4, 6, "Photo could not be embedded")

    row = max(row, 10)
    row = _section(sheet, row, "Dimension performance", 12)
    dimension_start = row + 1
    dimension_rows = []
    for code in DIMENSION_ORDER:
        item = (result or {}).get("dimensions", {}).get(code, {})
        dimension_rows.append([DIMENSION_NAMES[code], item.get("score", 0), item.get("rank", "—"), "Complete" if item.get("complete") else "Incomplete"])
    row = _write_table(sheet, row, ["Dimension", "Grade /1", "Rank", "Status"], dimension_rows, widths=[25, 14, 10, 16], number_formats={2: "0.00"})
    _add_radar(sheet, dimension_start, dimension_start + len(dimension_rows) - 1, 1, 2, "F11", "Dimension profile /1", 1)

    row = max(row, 23)
    row = _section(sheet, row, "Activity performance", 12)
    activity_start = row + 1
    activity_rows = []
    for code in ACTIVITY_ORDER:
        item = (result or {}).get("activities", {}).get(code, {})
        activity_rows.append([RUBRICS[code].name, item.get("score", 0), item.get("rank", "—"), item.get("submitted", 0), item.get("expected", 0), "Complete" if item.get("complete") else "Incomplete"])
    row = _write_table(sheet, row, ["Activity", "Grade /5", "Rank", "Submitted", "Expected", "Status"], activity_rows, widths=[22, 14, 10, 13, 13, 16], number_formats={2: "0.00"})
    _add_radar(sheet, activity_start, activity_start + len(activity_rows) - 1, 1, 2, "H24", "Activity profile /5", 5)

    row = max(row, 36)
    row = _section(sheet, row, "General assessment, comments and notes", 12)
    general_values = [
        ["Punctuality /1", float(assessment.punctuality) if assessment and assessment.punctuality is not None else "Missing"],
        ["Respect to us /1", float(assessment.respect) if assessment and assessment.respect is not None else "Missing"],
        ["Seriousness /1", float(assessment.seriousness) if assessment and assessment.seriousness is not None else "Missing"],
        ["General average /1", (result or {}).get("generalAverage", 0)],
    ]
    row = _write_table(sheet, row, ["General component", "Grade"], general_values, widths=[24, 16], number_formats={2: "0.00"})
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    sheet.cell(row, 1, f"General admin comment: {(assessment.comment if assessment else '') or 'None'}")
    sheet.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.cell(row, 1).fill = PatternFill("solid", fgColor=GRAY_LIGHT)
    sheet.row_dimensions[row].height = 34
    row += 1
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    sheet.cell(row, 1, f"Notes: {(assessment.notes if assessment else '') or 'None'}")
    sheet.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.cell(row, 1).fill = PatternFill("solid", fgColor=YELLOW_LIGHT)
    sheet.row_dimensions[row].height = 42
    row += 2

    row = _section(sheet, row, "Evaluator breakdown", 12)
    evaluator_rows = []
    criterion_rows = []
    for assignment in data["assignments"]:
        if assignment.recruit_id != recruit.id:
            continue
        code = data["assignment_activity"].get(assignment.id)
        evaluator = data["evaluator_by_id"].get(assignment.evaluator_id)
        submission = data["submissions"].get(assignment.id)
        evaluator_rows.append([RUBRICS[code].name, evaluator.name if evaluator else "Unknown", evaluator.role.title() if evaluator else "", float(submission.score) if submission else "", _label(submission.status) if submission else "Missing", submission.comments if submission else ""])
        if submission:
            responses = loads(submission.responses_json, {})
            raw = loads(submission.raw_payload_json, {})
            for criterion in RUBRICS[code].criteria:
                result_value = raw.get(criterion.key, "")
                criterion_rows.append([RUBRICS[code].name, criterion.dimension, criterion.name, evaluator.name if evaluator else "Unknown", responses.get(criterion.key, ""), f"{result_value} {criterion.unit}".strip() if result_value != "" else "", _label(submission.status)])
    for evaluation in data["admin_evaluations"]:
        if evaluation.recruit_id != recruit.id:
            continue
        responses = loads(evaluation.responses_json, {})
        raw = loads(evaluation.raw_payload_json, {})
        evaluator_rows.append([RUBRICS[evaluation.activity_code].name, evaluation.updated_by, "Admin", float(evaluation.score), "Official", evaluation.comments])
        for criterion in RUBRICS[evaluation.activity_code].criteria:
            result_value = raw.get(criterion.key, "")
            criterion_rows.append([RUBRICS[evaluation.activity_code].name, criterion.dimension, criterion.name, evaluation.updated_by, responses.get(criterion.key, ""), f"{result_value} {criterion.unit}".strip() if result_value != "" else "", "Official admin evaluation"])
    row = _write_table(sheet, row, ["Activity", "Evaluator", "Role", "Score /5", "Status", "Comment"], evaluator_rows, widths=[18, 26, 13, 13, 15, 42], number_formats={4: "0.00"})
    row = _section(sheet, row, "Criterion-level grading", 12)
    _write_table(sheet, row, ["Activity", "Dimension / theme", "Criterion", "Evaluator", "Grade /5", "Sport result", "Status"], criterion_rows, widths=[18, 19, 38, 25, 13, 16, 14], number_formats={5: "0.00"}, auto_filter=True)


def _audit_sheet(workbook: Workbook, journey: Journey, db: Session) -> None:
    sheet = _new_sheet(workbook, "Audit History", landscape=True)
    _title(sheet, "Administrative History", f"Readable change history · {journey.name}", 8)
    events = list(db.scalars(select(AuditEvent).where(AuditEvent.journey_id == journey.id).order_by(AuditEvent.created_at.desc())))
    rows = [[_local_time(item.created_at), item.actor_name, _label(item.actor_type), _label(item.action), _label(item.entity_type), item.entity_id or "", item.before_json or "", item.after_json or "", item.reason or ""] for item in events]
    _write_table(sheet, 4, ["Date and time (Beirut)", "Username", "Account type", "Action", "Location", "Record ID", "Before", "After", "Reason"], rows, widths=[25, 25, 15, 34, 20, 38, 55, 55, 45], auto_filter=True)


def build_report_workbook(db: Session, journey: Journey, *, full: bool) -> Workbook:
    del journey, full
    return build_management_report_workbook(db)


ALL_COMPLETED = "All completed Journees"


def _combined_results(journey_data: list[tuple[Journey, dict]]) -> dict:
    rows: list[dict] = []
    for journey, data in journey_data:
        for source in data["results"]["rows"]:
            row = deepcopy(source)
            row["journeyId"] = journey.id
            row["journeyName"] = journey.name
            row["profileKey"] = f"{journey.id}:{row['recruitId']}"
            rows.append(row)
    overall = competition_ranks([(row["profileKey"], Decimal(str(row["overallScore"]))) for row in rows])
    dimensions = {
        code: competition_ranks([(row["profileKey"], Decimal(str(row["dimensions"][code]["score"]))) for row in rows])
        for code in DIMENSION_ORDER
    }
    activities = {
        code: competition_ranks([(row["profileKey"], Decimal(str(row["activities"][code]["score"]))) for row in rows])
        for code in ACTIVITY_ORDER
    }
    for row in rows:
        row["overallRank"] = overall.get(row["profileKey"])
        for code in DIMENSION_ORDER:
            row["dimensions"][code]["rank"] = dimensions[code].get(row["profileKey"])
        for code in ACTIVITY_ORDER:
            row["activities"][code]["rank"] = activities[code].get(row["profileKey"])
    return {"rows": rows}


def _management_data(db: Session) -> tuple[list[Journey], dict[str, dict], dict]:
    journeys = list(db.scalars(select(Journey).order_by(Journey.event_date.desc(), Journey.name)))
    by_id: dict[str, dict] = {}
    completed: list[tuple[Journey, dict]] = []
    for journey in journeys:
        data = _collect(db, journey)
        data["db"] = db
        by_id[journey.id] = data
        if journey.status == "completed":
            completed.append((journey, data))
    return journeys, by_id, _combined_results(completed)


def _style_selector(sheet, label_cell: str, value_cell: str, label: str) -> None:
    sheet[label_cell] = label
    sheet[label_cell].font = Font(name="Aptos", size=9, bold=True, color=GRAY)
    sheet[value_cell].fill = PatternFill("solid", fgColor=WHITE)
    sheet[value_cell].font = Font(name="Aptos", size=10, bold=True, color=NAVY)
    sheet[value_cell].border = Border(left=THIN_LINE, right=THIN_LINE, top=THIN_LINE, bottom=THIN_LINE)
    sheet[value_cell].alignment = Alignment(vertical="center")


def _validation(sheet, cell: str, source_column: str, count: int) -> None:
    if count <= 0:
        return
    validation = DataValidation(type="list", formula1=f"=${source_column}$2:${source_column}${count + 1}", allow_blank=False)
    validation.error = "Choose a value from the list."
    validation.errorTitle = "Invalid selection"
    validation.prompt = "Select from the dropdown."
    validation.promptTitle = "Report selector"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)
    validation.add(sheet[cell])


def _write_hidden_rows(sheet, start_column: int, headers: list[str], rows: list[list]) -> tuple[int, int]:
    for offset, header in enumerate(headers):
        sheet.cell(1, start_column + offset, header)
        sheet.column_dimensions[get_column_letter(start_column + offset)].hidden = True
    for row_index, values in enumerate(rows, 2):
        for offset, value in enumerate(values):
            sheet.cell(row_index, start_column + offset, value)
    return 2, max(2, len(rows) + 1)


def _with_lookup_keys(rows: list[list], *key_indexes: int) -> list[list]:
    counters: defaultdict[tuple[str, ...], int] = defaultdict(int)
    keyed: list[list] = []
    for row in rows:
        group = tuple(str(row[index]) for index in key_indexes)
        counters[group] += 1
        keyed.append([*row, "|".join((*group, str(counters[group])))])
    return keyed


def _lookup_value_formula(
    *,
    key_expression: str,
    lookup_column: str,
    result_column: str,
    start: int,
    end: int,
) -> str:
    lookup = f'_xlfn.XLOOKUP({key_expression},${lookup_column}${start}:${lookup_column}${end},${result_column}${start}:${result_column}${end})'
    return f'=IFERROR(IF({lookup}="","",{lookup}),"")'


def _scalar_lookup_formula(key_expression: str, lookup_range: str, result_range: str, *, blank: str = "—") -> str:
    lookup = f"_xlfn.XLOOKUP({key_expression},{lookup_range},{result_range})"
    escaped_blank = blank.replace('"', '""')
    return f'=IFERROR(IF({lookup}="","{escaped_blank}",{lookup}),"{escaped_blank}")'


def _management_attendance_sheet(workbook: Workbook, db: Session, journeys: list[Journey], by_id: dict[str, dict]) -> None:
    sheet = _new_sheet(workbook, "Attendance", landscape=True)
    sheet.freeze_panes = "A6"
    _title(sheet, "Attendance", "Confirmed recruit and evaluator attendance", 9)
    _style_selector(sheet, "A3", "B3", "Journee view")
    _style_selector(sheet, "D3", "E3", "Roster")
    scopes = [ALL_COMPLETED, *[journey.name for journey in journeys]]
    sheet["B3"] = ALL_COMPLETED
    sheet["E3"] = "Recruits"
    rows: list[list] = []
    for journey in journeys:
        data = by_id[journey.id]
        record_scopes = [journey.name] + ([ALL_COMPLETED] if journey.status == "completed" else [])
        for scope in record_scopes:
            for recruit in data["recruits"]:
                rows.append([scope, "Recruits", journey.name, recruit.name, recruit.phone_number or "", recruit.date_of_birth, "Present" if recruit.present else "Absent", _local_time(recruit.arrival_time) if recruit.arrival_time else "", recruit.attendance_comment or "", "", ""])
            mandatory = {item.evaluator_id: item.room_number for item in db.scalars(select(MandatoryRoomEvaluator).where(MandatoryRoomEvaluator.journey_id == journey.id))}
            for evaluator in data["evaluators"]:
                rows.append([scope, "Evaluators", journey.name, evaluator.name, "", "", "Present" if evaluator.present else "Absent", "", "", evaluator.role.title(), f"Room {mandatory[evaluator.id]}" if evaluator.id in mandatory else ""])
    rows = _with_lookup_keys(rows, 0, 1)
    start, end = _write_hidden_rows(sheet, 13, ["Scope", "Roster", "Journee", "Name", "Phone number", "Date of birth", "Status", "Arrival time", "Attendance comment", "Role", "Mandatory room", "Lookup key"], rows)
    scope_col = 25
    for index, value in enumerate(scopes, 2):
        sheet.cell(index, scope_col, value)
    sheet.cell(1, scope_col, "Scope options")
    sheet.column_dimensions[get_column_letter(scope_col)].hidden = True
    sheet["Z1"] = "Roster options"; sheet["Z2"] = "Recruits"; sheet["Z3"] = "Evaluators"; sheet.column_dimensions["Z"].hidden = True
    _validation(sheet, "B3", "Y", len(scopes))
    _validation(sheet, "E3", "Z", 2)
    headers = ["Journee", "Name", "Phone number", "Date of birth", "Status", "Arrival time", "Attendance comment", "Role", "Mandatory room"]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(5, column, header); cell.fill = PatternFill("solid", fgColor=NAVY); cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    maximum_rows = max(Counter((row[0], row[1]) for row in rows).values(), default=1)
    for visible_row in range(6, 6 + maximum_rows):
        for column, source_column in enumerate(("O", "P", "Q", "R", "S", "T", "U", "V", "W"), 1):
            sheet.cell(visible_row, column, _lookup_value_formula(
                key_expression=f'$B$3&"|"&$E$3&"|"&ROWS($A$6:$A{visible_row})',
                lookup_column="X", result_column=source_column, start=start, end=end,
            ))
    for column, width in enumerate([27, 28, 18, 16, 12, 24, 34, 14, 18], 1): sheet.column_dimensions[get_column_letter(column)].width = width
    for visible_row in range(6, 6 + maximum_rows):
        sheet.cell(visible_row, 4).number_format = "dd mmm yyyy"
    sheet.auto_filter.ref = "A5:I5"


def _result_rows(scope: str, journey_name: str, results: dict) -> list[list]:
    rows: list[list] = []
    for item in results["rows"]:
        display_journey = item.get("journeyName", journey_name)
        rows.append([scope, "Overall ranking", item["overallRank"], item["name"], display_journey, item["overallScore"], "/20", f"{item['missingCount']} missing", "Complete" if item["complete"] else "Incomplete", item["color"].title()])
        for code in DIMENSION_ORDER:
            value = item["dimensions"][code]
            rows.append([scope, DIMENSION_NAMES[code], value["rank"], item["name"], display_journey, value["score"], "/1", f"{round(value.get('availableWeight', 0) * 100)}% coverage", "Complete" if value["complete"] else "Incomplete", ""])
        for code in ACTIVITY_ORDER:
            value = item["activities"][code]
            rows.append([scope, RUBRICS[code].name, value["rank"], item["name"], display_journey, value["score"], "/5", f"{value['submitted']}/{value['expected']} submitted", "Complete" if value["complete"] else "Incomplete", ""])
    return rows


def _management_results_sheet(workbook: Workbook, journeys: list[Journey], by_id: dict[str, dict], combined: dict) -> None:
    sheet = _new_sheet(workbook, "Results", landscape=True)
    sheet.freeze_panes = "A6"
    _title(sheet, "Results & rankings", "Overall, dimension, and activity rankings", 8)
    _style_selector(sheet, "A3", "B3", "Journee view")
    _style_selector(sheet, "D3", "E3", "Result view")
    scopes = [ALL_COMPLETED, *[journey.name for journey in journeys]]
    views = ["Overall ranking", *[DIMENSION_NAMES[code] for code in DIMENSION_ORDER], *[RUBRICS[code].name for code in ACTIVITY_ORDER]]
    sheet["B3"] = ALL_COMPLETED; sheet["E3"] = "Overall ranking"
    rows = _result_rows(ALL_COMPLETED, "", combined)
    for journey in journeys:
        rows.extend(_result_rows(journey.name, journey.name, by_id[journey.id]["results"]))
    rows = _with_lookup_keys(rows, 0, 1)
    start, end = _write_hidden_rows(sheet, 13, ["Scope", "View", "Rank", "Recruit", "Journee", "Score", "Scale", "Details", "Status", "Color", "Lookup key"], rows)
    for index, value in enumerate(scopes, 2): sheet.cell(index, 24, value)
    for index, value in enumerate(views, 2): sheet.cell(index, 25, value)
    sheet.cell(1, 24, "Scope options"); sheet.cell(1, 25, "View options")
    sheet.column_dimensions["X"].hidden = True; sheet.column_dimensions["Y"].hidden = True
    _validation(sheet, "B3", "X", len(scopes)); _validation(sheet, "E3", "Y", len(views))
    headers = ["Rank", "Recruit", "Journee", "Score", "Scale", "Details", "Status", "Color"]
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(5, column, header); cell.fill = PatternFill("solid", fgColor=NAVY); cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    maximum_rows = max(Counter((row[0], row[1]) for row in rows).values(), default=1)
    for visible_row in range(6, 6 + maximum_rows):
        for column, source_column in enumerate(("O", "P", "Q", "R", "S", "T", "U", "V"), 1):
            sheet.cell(visible_row, column, _lookup_value_formula(
                key_expression=f'$B$3&"|"&$E$3&"|"&ROWS($A$6:$A{visible_row})',
                lookup_column="W", result_column=source_column, start=start, end=end,
            ))
    for column, width in enumerate([10, 30, 27, 14, 10, 23, 15, 12], 1): sheet.column_dimensions[get_column_letter(column)].width = width
    for visible_row in range(6, 6 + maximum_rows):
        sheet.cell(visible_row, 4).number_format = "0.00"


def _fallback_result(recruit: Recruit) -> dict:
    return {
        "overallScore": 0.0, "overallRank": None, "color": "red", "missingCount": len(ACTIVITY_ORDER) + 3,
        "missingComponents": [RUBRICS[code].name for code in ACTIVITY_ORDER] + ["Punctuality", "Respect to us", "Seriousness"],
        "dimensions": {code: {"score": 0.0, "rank": None, "complete": False, "availableWeight": 0.0} for code in DIMENSION_ORDER},
        "activities": {code: {"score": 0.0, "rank": None, "complete": False, "submitted": 0, "expected": 0} for code in ACTIVITY_ORDER},
        "generalAverage": 0.0, "complete": False, "name": recruit.name,
    }


def _profile_label(recruit: Recruit, journey: Journey, duplicates: dict[tuple[str, str], int]) -> str:
    suffix = f" · {recruit.id[:6]}" if duplicates[(journey.id, recruit.name.casefold())] > 1 else ""
    return f"{recruit.name}{suffix} · {journey.name}"


def _management_profile_sheet(workbook: Workbook, db: Session, journeys: list[Journey], by_id: dict[str, dict], combined: dict) -> None:
    sheet = _new_sheet(workbook, "Recruit Profiles", landscape=True)
    sheet.freeze_panes = "A7"
    _title(sheet, "Recruit profile", "Complete view-only recruit record", 12)
    _style_selector(sheet, "A3", "B3", "Journee view")
    _style_selector(sheet, "D3", "E3", "Recruit")
    duplicates: defaultdict[tuple[str, str], int] = defaultdict(int)
    for journey in journeys:
        for recruit in by_id[journey.id]["recruits"]: duplicates[(journey.id, recruit.name.casefold())] += 1
    combined_by_key = {row["profileKey"]: row for row in combined["rows"]}
    summary_rows: list[list] = []; dimension_rows: list[list] = []; activity_rows: list[list] = []
    evaluator_rows: list[list] = []; criterion_rows: list[list] = []; audit_rows: list[list] = []
    selector_labels: list[str] = []
    scope_names = [ALL_COMPLETED, *[journey.name for journey in journeys]]
    for journey in journeys:
        data = by_id[journey.id]
        for recruit in data["recruits"]:
            profile_key = f"{journey.id}:{recruit.id}"
            label = _profile_label(recruit, journey, duplicates)
            selector_labels.append(label)
            base_result = data["result_by_recruit"].get(recruit.id) or _fallback_result(recruit)
            scoped = [(journey.name, base_result)]
            if journey.status == "completed": scoped.append((ALL_COMPLETED, combined_by_key.get(profile_key, base_result)))
            assessment = data["assessments"].get(recruit.id)
            for scope, result in scoped:
                selection_key = f"{scope}|{label}"
                summary_rows.append([selection_key, profile_key, journey.name, journey.event_date, recruit.name, recruit.phone_number or "", recruit.date_of_birth, "Present" if recruit.present else "Absent", _local_time(recruit.arrival_time) if recruit.arrival_time else "", recruit.attendance_comment or "", result["overallScore"], result["overallRank"] or "", result["color"].title(), ", ".join(result.get("missingComponents", [])) or "Complete", float(assessment.punctuality) if assessment and assessment.punctuality is not None else "", float(assessment.respect) if assessment and assessment.respect is not None else "", float(assessment.seriousness) if assessment and assessment.seriousness is not None else "", float(result.get("generalAverage", 0)), assessment.comment if assessment else "", assessment.notes if assessment else ""])
                for code in DIMENSION_ORDER:
                    value = result["dimensions"][code]
                    dimension_rows.append([selection_key, DIMENSION_NAMES[code], value["score"], value["rank"] or "", "Complete" if value["complete"] else "Incomplete", f"{round(value.get('availableWeight', 0) * 100)}%"])
                for code in ACTIVITY_ORDER:
                    value = result["activities"][code]
                    activity_rows.append([selection_key, RUBRICS[code].name, value["score"], value["rank"] or "", f"{value['submitted']}/{value['expected']}", "Complete" if value["complete"] else "Incomplete"])
            for assignment in data["assignments"]:
                if assignment.recruit_id != recruit.id: continue
                code = data["assignment_activity"].get(assignment.id)
                evaluator = data["evaluator_by_id"].get(assignment.evaluator_id)
                submission = data["submissions"].get(assignment.id)
                evaluator_rows.append([profile_key, RUBRICS[code].name if code else "", evaluator.name if evaluator else "Unknown", evaluator.role.title() if evaluator else "", float(submission.score) if submission else "", _label(submission.status) if submission else "Missing", submission.comments if submission else ""])
                if code and submission:
                    responses = loads(submission.responses_json, {}); raw = loads(submission.raw_payload_json, {})
                    for criterion in RUBRICS[code].criteria:
                        raw_value = raw.get(criterion.key, "")
                        criterion_rows.append([profile_key, RUBRICS[code].name, criterion.dimension, criterion.name, criterion.explanation, evaluator.name if evaluator else "Unknown", responses.get(criterion.key, ""), f"{raw_value} {criterion.unit}".strip() if raw_value != "" else "", _label(submission.status)])
            for evaluation in data["admin_evaluations"]:
                if evaluation.recruit_id != recruit.id: continue
                evaluator_rows.append([profile_key, RUBRICS[evaluation.activity_code].name, f"Admin: {evaluation.updated_by}", "Admin", float(evaluation.score), "Official", evaluation.comments])
                responses = loads(evaluation.responses_json, {}); raw = loads(evaluation.raw_payload_json, {})
                for criterion in RUBRICS[evaluation.activity_code].criteria:
                    raw_value = raw.get(criterion.key, "")
                    criterion_rows.append([profile_key, RUBRICS[evaluation.activity_code].name, criterion.dimension, criterion.name, criterion.explanation, f"Admin: {evaluation.updated_by}", responses.get(criterion.key, ""), f"{raw_value} {criterion.unit}".strip() if raw_value != "" else "", "Official admin evaluation"])
            for event in db.scalars(select(AuditEvent).where(AuditEvent.journey_id == journey.id, AuditEvent.entity_id == recruit.id).order_by(AuditEvent.created_at.desc())):
                audit_rows.append([profile_key, _local_time(event.created_at), event.actor_name, _label(event.action), event.reason or "", event.before_json or "", event.after_json or ""])
    selector_labels = sorted(set(selector_labels), key=str.casefold)
    default_label = next((label for label in selector_labels if any(row[0] == f"{ALL_COMPLETED}|{label}" for row in summary_rows)), selector_labels[0] if selector_labels else "")
    sheet["B3"] = ALL_COMPLETED if any(j.status == "completed" for j in journeys) else (journeys[0].name if journeys else "")
    sheet["E3"] = default_label
    _write_hidden_rows(sheet, 27, ["Selection", "Profile key", "Journee", "Date", "Recruit", "Phone", "DOB", "Attendance", "Arrival", "Attendance comment", "Overall", "Rank", "Color", "Missing", "Punctuality", "Respect", "Seriousness", "General average", "General comment", "Notes"], summary_rows)
    _write_hidden_rows(sheet, 49, ["Selection", "Dimension", "Score", "Rank", "Status", "Coverage"], dimension_rows)
    _write_hidden_rows(sheet, 56, ["Selection", "Activity", "Score", "Rank", "Submissions", "Status"], activity_rows)
    evaluator_rows = _with_lookup_keys(evaluator_rows, 0)
    criterion_rows = _with_lookup_keys(criterion_rows, 0)
    audit_rows = _with_lookup_keys(audit_rows, 0)
    _write_hidden_rows(sheet, 63, ["Profile key", "Activity", "Evaluator", "Role", "Score", "Status", "Comment", "Lookup key"], evaluator_rows)
    _write_hidden_rows(sheet, 71, ["Profile key", "Activity", "Dimension", "Criterion", "Explanation", "Evaluator", "Grade", "Sport result", "Status", "Lookup key"], criterion_rows)
    _write_hidden_rows(sheet, 81, ["Profile key", "Date", "Username", "Action", "Reason", "Before", "After", "Lookup key"], audit_rows)
    for index, value in enumerate(scope_names, 2): sheet.cell(index, 89, value)
    for index, value in enumerate(selector_labels, 2): sheet.cell(index, 90, value)
    sheet.column_dimensions["CK"].hidden = True; sheet.column_dimensions["CL"].hidden = True
    _validation(sheet, "B3", "CK", len(scope_names)); _validation(sheet, "E3", "CL", len(selector_labels))
    last_summary = max(2, len(summary_rows) + 1); last_dimension = max(2, len(dimension_rows) + 1); last_activity = max(2, len(activity_rows) + 1)
    selection = '$B$3&"|"&$E$3'
    sheet["H3"] = _scalar_lookup_formula(selection, f"$AA$2:$AA${last_summary}", f"$AB$2:$AB${last_summary}", blank="")
    sheet["H3"].number_format = ";;;"
    fields = [("A5", "Recruit", "AE"), ("D5", "Journee", "AC"), ("G5", "Date", "AD"), ("J5", "Attendance", "AH"), ("A6", "Phone", "AF"), ("D6", "Date of birth", "AG"), ("G6", "Arrival", "AI"), ("J6", "Overall", "AK")]
    for cell, label, source in fields:
        sheet[cell] = f'{label}: '; sheet[cell].font = Font(name="Aptos", size=9, bold=True, color=GRAY)
        value_cell = sheet.cell(sheet[cell].row, sheet[cell].column + 1)
        value_cell.value = _scalar_lookup_formula(selection, f"$AA$2:$AA${last_summary}", f"${source}$2:${source}${last_summary}")
        value_cell.font = Font(name="Aptos", size=11, bold=True, color=NAVY)
    sheet["E6"].number_format = "dd mmm yyyy"
    sheet["H5"].number_format = "dd mmm yyyy"
    sheet["K6"] = f'=IFERROR(TEXT(_xlfn.XLOOKUP({selection},$AA$2:$AA${last_summary},$AK$2:$AK${last_summary}),"0.00")&" /20 · rank "&_xlfn.XLOOKUP({selection},$AA$2:$AA${last_summary},$AL$2:$AL${last_summary}),"—")'
    row = _section(sheet, 8, "Dimension performance", 12)
    for col, header in enumerate(["Dimension", "Score /1", "Rank", "Status", "Coverage"], 1):
        sheet.cell(row, col, header).fill = PatternFill("solid", fgColor=NAVY); sheet.cell(row, col).font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    for offset, code in enumerate(DIMENSION_ORDER, 1):
        target = row + offset; sheet.cell(target, 1, DIMENSION_NAMES[code])
        for col, source_col in enumerate(["AY", "AZ", "BA", "BB"], 2):
            sheet.cell(target, col, _scalar_lookup_formula(
                f'{selection}&"|"&$A{target}',
                f'$AW$2:$AW${last_dimension}&"|"&$AX$2:$AX${last_dimension}',
                f'${source_col}$2:${source_col}${last_dimension}',
            ))
    _add_radar(sheet, row + 1, row + len(DIMENSION_ORDER), 1, 2, "G8", "Dimension performance", 1)
    activity_section = 18
    row = _section(sheet, activity_section, "Activity performance", 12)
    for col, header in enumerate(["Activity", "Score /5", "Rank", "Submissions", "Status"], 1):
        sheet.cell(row, col, header).fill = PatternFill("solid", fgColor=NAVY); sheet.cell(row, col).font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    for offset, code in enumerate(ACTIVITY_ORDER, 1):
        target = row + offset; sheet.cell(target, 1, RUBRICS[code].name)
        for col, source_col in enumerate(["BF", "BG", "BH", "BI"], 2):
            sheet.cell(target, col, _scalar_lookup_formula(
                f'{selection}&"|"&$A{target}',
                f'$BD$2:$BD${last_activity}&"|"&$BE$2:$BE${last_activity}',
                f'${source_col}$2:${source_col}${last_activity}',
            ))
    _add_radar(sheet, row + 1, row + len(ACTIVITY_ORDER), 1, 2, "G18", "Activity performance", 5)
    row = _section(sheet, 28, "General assessment and completion", 12)
    labels = [("Punctuality /1", "AO"), ("Respect to us /1", "AP"), ("Seriousness /1", "AQ"), ("General average /1", "AR"), ("Color grade", "AM"), ("Missing components", "AN"), ("General comment", "AS"), ("Notes", "AT")]
    for index, (label, source) in enumerate(labels):
        target = row + index; sheet.cell(target, 1, label); sheet.cell(target, 1).font = Font(name="Aptos", size=9, bold=True, color=GRAY)
        sheet.merge_cells(start_row=target, start_column=2, end_row=target, end_column=12)
        sheet.cell(target, 2, _scalar_lookup_formula(selection, f"$AA$2:$AA${last_summary}", f"${source}$2:${source}${last_summary}"))
        sheet.cell(target, 2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    row = _section(sheet, 39, "Evaluator breakdown", 12)
    evaluator_headers = ["Activity", "Evaluator", "Role", "Score /5", "Status", "Comment"]
    for col, header in enumerate(evaluator_headers, 1): sheet.cell(row, col, header).fill = PatternFill("solid", fgColor=NAVY); sheet.cell(row, col).font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    evaluator_end = max(2, len(evaluator_rows) + 1)
    evaluator_visible_start = row + 1
    evaluator_maximum = max(Counter(item[0] for item in evaluator_rows).values(), default=1)
    for visible_row in range(evaluator_visible_start, evaluator_visible_start + evaluator_maximum):
        for column, source_column in enumerate(("BL", "BM", "BN", "BO", "BP", "BQ"), 1):
            sheet.cell(visible_row, column, _lookup_value_formula(
                key_expression=f'$H$3&"|"&ROWS($A${evaluator_visible_start}:$A{visible_row})',
                lookup_column="BR", result_column=source_column, start=2, end=evaluator_end,
            ))
    row = _section(sheet, 61, "Criterion-level grading", 12)
    criterion_headers = ["Activity", "Dimension", "Criterion", "Explanation", "Evaluator", "Grade /5", "Sport result", "Status"]
    for col, header in enumerate(criterion_headers, 1): sheet.cell(row, col, header).fill = PatternFill("solid", fgColor=NAVY); sheet.cell(row, col).font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    criterion_end = max(2, len(criterion_rows) + 1)
    criterion_visible_start = row + 1
    criterion_maximum = min(116, max(Counter(item[0] for item in criterion_rows).values(), default=1))
    for visible_row in range(criterion_visible_start, criterion_visible_start + criterion_maximum):
        for column, source_column in enumerate(("BT", "BU", "BV", "BW", "BX", "BY", "BZ", "CA"), 1):
            sheet.cell(visible_row, column, _lookup_value_formula(
                key_expression=f'$H$3&"|"&ROWS($A${criterion_visible_start}:$A{visible_row})',
                lookup_column="CB", result_column=source_column, start=2, end=criterion_end,
            ))
    row = _section(sheet, 180, "Profile audit history", 12)
    audit_headers = ["Date and time", "Username", "Action", "Reason", "Before", "After"]
    for col, header in enumerate(audit_headers, 1): sheet.cell(row, col, header).fill = PatternFill("solid", fgColor=NAVY); sheet.cell(row, col).font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    audit_end = max(2, len(audit_rows) + 1)
    audit_visible_start = row + 1
    audit_maximum = max(Counter(item[0] for item in audit_rows).values(), default=1)
    for visible_row in range(audit_visible_start, audit_visible_start + audit_maximum):
        for column, source_column in enumerate(("CD", "CE", "CF", "CG", "CH", "CI"), 1):
            sheet.cell(visible_row, column, _lookup_value_formula(
                key_expression=f'$H$3&"|"&ROWS($A${audit_visible_start}:$A{visible_row})',
                lookup_column="CJ", result_column=source_column, start=2, end=audit_end,
            ))
    for col, width in enumerate([24, 17, 12, 34, 24, 16, 19, 16, 16, 16, 16, 16], 1): sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.row_dimensions[5].height = 24; sheet.row_dimensions[6].height = 30


def build_management_report_workbook(db: Session) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    journeys, by_id, combined = _management_data(db)
    _management_attendance_sheet(workbook, db, journeys, by_id)
    _management_results_sheet(workbook, journeys, by_id, combined)
    _management_profile_sheet(workbook, db, journeys, by_id, combined)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.calculation.calcId = 0
    return workbook
