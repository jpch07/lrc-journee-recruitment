from __future__ import annotations

import csv
import io
import json
import random
import re
import secrets
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
import qrcode
from sqlalchemy import delete, func, select, update
from sqlalchemy.orm import Session

from .auth import (
    AdminContext,
    clear_expired_sessions,
    clear_login_attempts,
    create_admin_session,
    enforce_login_rate_limit,
    logout_admin,
    require_admin,
    require_csrf,
    verify_admin_password,
    hash_password,
)
from .db import get_db
from .config import settings
from .event_day import ProtectionControlError, cancel_monitor, dispatch_monitor, find_monitor_run, is_configured
from .models import (
    ActivityState,
    AdminEvaluation,
    Assignment,
    AssignmentRound,
    AuditEvent,
    EvaluationSubmission,
    Evaluator,
    EvaluatorDirectory,
    EventDayProtection,
    GeneralAssessment,
    IdempotencyRecord,
    Journey,
    MandatoryRoomEvaluator,
    Recruit,
    RecruitAttendanceAccess,
    RecruitAttendanceSession,
    RoomPlan,
    RoomPlanEvaluator,
    RoomPlanRecruit,
    SubmissionVersion,
    UserAccount,
    new_id,
    utcnow,
)
from .rubric import (
    ACTIVITY_ORDER,
    BEHAVIORAL_DIMENSIONS,
    DIMENSION_ACTIVITIES,
    DIMENSION_NAMES,
    DIMENSION_ORDER,
    ROOM_ACTIVITIES,
    RUBRICS,
    public_rubric,
)
from .report_exports import build_report_workbook
from .schemas import (
    ActivityActionRequest,
    AdminCorrectionRequest,
    AdminEvaluationRequest,
    AdminLoginRequest,
    AssignmentEditRequest,
    EvaluatorAttendanceRequest,
    EvaluationSimulationRequest,
    EventDayProtectionActionRequest,
    EventDayProtectionStartRequest,
    EvaluatorCreateRequest,
    GeneralAssessmentRequest,
    JourneyCreateRequest,
    JourneyUpdateRequest,
    MandatoryRoomRequest,
    PreviewRequest,
    RecruitAttendanceRequest,
    RecruitAttendancePatchRequest,
    RecruitCreateRequest,
    RoomCountRequest,
    RoomPlanEditRequest,
)
from .scoring import ScoringError, validate_general_grade
from .services import (
    assignment_round_payload,
    create_assignment_preview,
    create_journey,
    create_room_preview,
    get_evaluator_or_404,
    get_journey_or_404,
    get_recruit_or_404,
    latest_room_plan,
    monitoring_snapshot,
    past_pair_data,
    process_photo,
    publish_assignment_round,
    publish_room_plan,
    result_snapshot,
    room_plan_payload,
    save_submission,
    score_evaluation,
    serialize_evaluator,
    serialize_journey,
    serialize_recruit,
    update_recruit_attendance_fields,
)
from .utils import audit, dumps, loads


router = APIRouter(prefix="/api/admin", tags=["admin"])


def _commit(db: Session) -> None:
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise


def _activity_states(db: Session, journey_id: str) -> list[dict]:
    states = {
        item.code: item
        for item in db.scalars(select(ActivityState).where(ActivityState.journey_id == journey_id))
    }
    return [
        {
            "code": code,
            "name": RUBRICS[code].name,
            "status": states[code].status,
            "version": states[code].version,
            "assignmentRoundId": states[code].assignment_round_id,
            "openedAt": states[code].opened_at.isoformat() if states[code].opened_at else None,
            "closedAt": states[code].closed_at.isoformat() if states[code].closed_at else None,
        }
        for code in ACTIVITY_ORDER
        if code in states
    ]


def _aware(value: datetime) -> datetime:
    return value if value.tzinfo else value.replace(tzinfo=timezone.utc)


def _sync_protection(db: Session, protection: EventDayProtection) -> None:
    if protection.status in {"stopped", "completed"}:
        return
    now = utcnow()
    if _aware(protection.ends_at) <= now:
        protection.status = "completed"
        protection.last_error = ""
        protection.version += 1
        return
    try:
        run = find_monitor_run(protection.activation_id)
    except ProtectionControlError as exc:
        protection.last_error = str(exc)
        protection.version += 1
        return
    if not run:
        if now - _aware(protection.started_at) > timedelta(minutes=2):
            protection.status = "incident"
            protection.last_error = "GitHub did not create the protection run within two minutes. Restart protection."
            protection.version += 1
        return
    changed = False
    run_id = str(run["id"])
    if protection.github_run_id != run_id:
        protection.github_run_id = run_id
        protection.github_run_url = run.get("html_url")
        changed = True
    failed_jobs = [
        job.get("name", "monitor")
        for job in run.get("jobs", [])
        if job.get("conclusion") in {"failure", "timed_out", "startup_failure"}
    ]
    if failed_jobs:
        status = "incident"
        error = f"A protection monitor failed: {', '.join(failed_jobs)}. Start replacement protection."
    elif run.get("status") in {"queued", "in_progress", "waiting", "pending"}:
        status = "active"
        error = ""
    elif run.get("conclusion") == "success":
        status = "completed"
        error = ""
    elif run.get("conclusion") == "cancelled" and protection.status == "stopped":
        status = "stopped"
        error = ""
    else:
        status = "incident"
        error = f"The protection workflow ended with status: {run.get('conclusion') or run.get('status')}."
    if protection.status != status or protection.last_error != error:
        protection.status = status
        protection.last_error = error
        changed = True
    if changed:
        protection.version += 1


def _protection_payload(protection: EventDayProtection | None) -> dict:
    if not protection:
        return {"configured": is_configured(), "status": "inactive"}
    now = utcnow()
    remaining = max(0, int((_aware(protection.ends_at) - now).total_seconds()))
    return {
        "configured": is_configured(),
        "status": protection.status,
        "durationHours": protection.duration_hours,
        "startedAt": _aware(protection.started_at).isoformat(),
        "endsAt": _aware(protection.ends_at).isoformat(),
        "remainingSeconds": remaining,
        "runUrl": protection.github_run_url,
        "error": protection.last_error,
        "version": protection.version,
    }


def _cancel_protection_monitor(protection: EventDayProtection) -> None:
    run_id = protection.github_run_id
    if not run_id:
        try:
            run = find_monitor_run(protection.activation_id)
            run_id = str(run["id"]) if run else None
        except ProtectionControlError:
            run_id = None
    if run_id:
        cancel_monitor(run_id)


@router.post("/login")
def admin_login(payload: AdminLoginRequest, request: Request, response: Response, db: Session = Depends(get_db)):
    enforce_login_rate_limit(request)
    if not verify_admin_password(payload.password):
        raise HTTPException(status_code=403, detail="Invalid admin password.")
    clear_login_attempts(request)
    clear_expired_sessions(db)
    session = create_admin_session(db, response, payload.display_name)
    _commit(db)
    return {"ok": True, "displayName": session.actor_name, "csrfToken": session.csrf_token,
            "testToolsEnabled": settings.allow_test_tools}


@router.get("/session")
def admin_session(context: AdminContext = Depends(require_admin)):
    return {"authenticated": True, "displayName": context.actor_name, "csrfToken": context.csrf_token,
            "testToolsEnabled": settings.allow_test_tools}


@router.post("/logout")
def admin_logout(
    request: Request,
    response: Response,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    logout_admin(db, response, context)
    _commit(db)
    return {"ok": True}


@router.get("/evaluator-directory")
def evaluator_directory(
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    return [
        {"id": item.id, "name": item.name, "defaultRole": item.default_role}
        for item in db.scalars(
            select(EvaluatorDirectory)
            .where(EvaluatorDirectory.active.is_(True))
            .order_by(EvaluatorDirectory.name)
        )
    ]


@router.post("/journeys/{journey_id}/evaluators/from-directory/{directory_id}")
def reuse_directory_evaluator(
    journey_id: str,
    directory_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    directory = db.get(EvaluatorDirectory, directory_id)
    if not directory or not directory.active:
        raise HTTPException(status_code=404, detail="Evaluator directory entry not found.")
    existing = db.scalar(
        select(Evaluator).where(
            Evaluator.journey_id == journey.id,
            Evaluator.directory_id == directory.id,
        )
    )
    if existing:
        if existing.active:
            raise HTTPException(status_code=409, detail="This evaluator is already in the Journee.")
        existing.active = True
        existing.version += 1
        evaluator = existing
    else:
        evaluator = Evaluator(
            journey_id=journey.id,
            directory_id=directory.id,
            name=directory.name,
            role=directory.default_role,
        )
        db.add(evaluator)
        db.flush()
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="evaluator.reused_from_directory",
        entity_type="evaluator",
        entity_id=evaluator.id,
    )
    _commit(db)
    return serialize_evaluator(evaluator)


def _import_rows(file: UploadFile) -> list[dict[str, str]]:
    raw = file.file.read()
    suffix = (file.filename or "").lower().rsplit(".", 1)[-1]
    if suffix in {"xlsx", "xlsm"}:
        try:
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
        except Exception as exc:
            raise HTTPException(status_code=400, detail="The uploaded Excel workbook could not be read.") from exc
        if not values:
            return []
        headers = [str(value or "").strip().lower() for value in values[0]]
        return [
            {headers[index]: str(value or "").strip() for index, value in enumerate(row) if index < len(headers)}
            for row in values[1:]
            if any(value not in {None, ""} for value in row)
        ]
    try:
        decoded = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail="CSV files must use UTF-8 encoding.") from exc
    return [
        {str(key or "").strip().lower(): str(value or "").strip() for key, value in row.items()}
        for row in csv.DictReader(io.StringIO(decoded))
    ]


def _first_value(row: dict[str, str], aliases: tuple[str, ...]) -> str:
    for alias in aliases:
        if row.get(alias, "").strip():
            return row[alias].strip()
    return ""


def _import_date_of_birth(value: str) -> date | None:
    raw = value.strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw).date()
    except ValueError:
        pass
    for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            continue
    raise ValueError("use YYYY-MM-DD or DD/MM/YYYY")


@router.post("/journeys/{journey_id}/import/{kind}")
def import_people(
    journey_id: str,
    kind: str,
    request: Request,
    file: UploadFile = File(...),
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    if kind not in {"recruits", "evaluators"}:
        raise HTTPException(status_code=404, detail="Import type must be recruits or evaluators.")
    rows = _import_rows(file)
    created: list[dict] = []
    skipped: list[str] = []
    for number, row in enumerate(rows, start=2):
        name = _first_value(row, ("name", "full name", "recruit", "recruit name", "evaluator", "evaluator name"))
        if not name:
            skipped.append(f"Row {number}: missing name")
            continue
        clean_name = " ".join(name.split())
        if kind == "recruits":
            phone_number = _first_value(row, ("phone", "phone number", "mobile", "mobile number")) or None
            date_of_birth_text = _first_value(row, ("date of birth", "dob", "birth date", "birthday"))
            try:
                date_of_birth = _import_date_of_birth(date_of_birth_text)
            except ValueError as exc:
                skipped.append(f"Row {number}: invalid date of birth ({exc})")
                continue
            duplicate = db.scalar(
                select(Recruit).where(
                    Recruit.journey_id == journey.id,
                    func.lower(Recruit.name) == clean_name.lower(),
                )
            )
            if duplicate:
                skipped.append(f"Row {number}: {clean_name} already exists")
                continue
            item = Recruit(
                journey_id=journey.id,
                name=clean_name,
                phone_number=phone_number,
                date_of_birth=date_of_birth,
            )
            db.add(item)
            db.flush()
            created.append(serialize_recruit(item))
        else:
            role_text = _first_value(row, ("role", "type")).lower()
            role = "overall" if role_text in {"overall", "o", "general"} else "dossard"
            duplicate = db.scalar(
                select(Evaluator).where(
                    Evaluator.journey_id == journey.id,
                    func.lower(Evaluator.name) == clean_name.lower(),
                )
            )
            if duplicate:
                skipped.append(f"Row {number}: {clean_name} already exists")
                continue
            directory = db.scalar(
                select(EvaluatorDirectory).where(func.lower(EvaluatorDirectory.name) == clean_name.lower())
            )
            if not directory:
                directory = EvaluatorDirectory(name=clean_name, default_role=role)
                db.add(directory)
                db.flush()
            item = Evaluator(journey_id=journey.id, directory_id=directory.id, name=clean_name, role=role)
            db.add(item)
            db.flush()
            created.append(serialize_evaluator(item))
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action=f"{kind}.imported",
        entity_type=kind,
        after={"created": len(created), "skipped": skipped, "filename": file.filename},
    )
    _commit(db)
    return {"created": created, "skipped": skipped}


def _filename_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


@router.post("/journeys/{journey_id}/recruits/photos")
def bulk_recruit_photos(
    journey_id: str,
    request: Request,
    photos: list[UploadFile] = File(...),
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    recruits = list(db.scalars(select(Recruit).where(Recruit.journey_id == journey.id, Recruit.active.is_(True))))
    keys: dict[str, list[Recruit]] = defaultdict(list)
    for recruit in recruits:
        keys[_filename_key(recruit.name)].append(recruit)
    matched: list[str] = []
    skipped: list[str] = []
    for photo in photos:
        stem = (photo.filename or "").rsplit(".", 1)[0]
        candidates = keys.get(_filename_key(stem), [])
        unique = {item.id: item for item in candidates}
        if len(unique) != 1:
            skipped.append(f"{photo.filename}: {'no unique name match' if unique else 'no match'}")
            continue
        recruit = next(iter(unique.values()))
        data, media_type = process_photo(photo.file.read())
        recruit.photo_data = data
        recruit.photo_type = media_type
        recruit.photo_updated_at = utcnow()
        recruit.version += 1
        matched.append(recruit.name)
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="recruit.photos_bulk_uploaded",
        entity_type="recruit_photo",
        after={"matched": matched, "skipped": skipped},
    )
    _commit(db)
    return {"matched": matched, "skipped": skipped}


@router.get("/journeys/{journey_id}/evaluator-qr.png")
def evaluator_qr(
    journey_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    journey = get_journey_or_404(db, journey_id)
    url = str(request.base_url).rstrip("/") + "/evaluate"
    image = qrcode.make(url)
    output = io.BytesIO()
    image.save(output, format="PNG")
    return Response(
        content=output.getvalue(),
        media_type="image/png",
        headers={"Cache-Control": "private, no-store"},
    )


@router.get("/journeys/{journey_id}/recruit-attendance-qr.png")
def recruit_attendance_qr(
    journey_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    journey = get_journey_or_404(db, journey_id)
    access = db.get(RecruitAttendanceAccess, journey.id)
    if not access:
        raise HTTPException(status_code=404, detail="Recruit attendance link is not configured.")
    url = str(request.base_url).rstrip("/") + f"/recruit-attendance/{access.token}"
    image = qrcode.make(url)
    output = io.BytesIO()
    image.save(output, format="PNG")
    return Response(
        content=output.getvalue(),
        media_type="image/png",
        headers={"Cache-Control": "private, no-store"},
    )


def _append_sheet(workbook: Workbook, name: str, headers: list[str], rows: list[list]) -> None:
    sheet = workbook.create_sheet(title=name[:31])
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 55)


def _full_export(db: Session, journey: Journey) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    recruits = list(db.scalars(select(Recruit).where(Recruit.journey_id == journey.id).order_by(Recruit.name)))
    evaluators = list(db.scalars(select(Evaluator).where(Evaluator.journey_id == journey.id).order_by(Evaluator.name)))
    recruit_by_id = {item.id: item for item in recruits}
    evaluator_by_id = {item.id: item for item in evaluators}
    _append_sheet(
        workbook,
        "Journee",
        ["Name", "Date", "Status", "Rooms", "Current activity", "Exported UTC"],
        [[journey.name, journey.event_date.isoformat(), journey.status, journey.room_count, journey.current_activity or "", utcnow().isoformat()]],
    )
    _append_sheet(
        workbook,
        "Rubric 2026",
        ["Activity", "Dimension", "Criterion", "Weight", "Explanation", "Input type", "Target", "Unit"],
        [
            [
                rubric.name,
                criterion.dimension,
                criterion.name,
                float(criterion.weight),
                criterion.explanation,
                criterion.input_type,
                float(criterion.target) if criterion.target is not None else "",
                criterion.unit,
            ]
            for rubric in RUBRICS.values()
            for criterion in rubric.criteria
        ],
    )
    _append_sheet(
        workbook,
        "Recruits",
        ["Recruit UUID", "Name", "Phone Number", "Date of Birth", "Active", "Present", "Arrival UTC", "Attendance Comment", "Has photo"],
        [[item.id, item.name, item.phone_number or "", item.date_of_birth.isoformat() if item.date_of_birth else "", item.active, item.present, item.arrival_time.isoformat() if item.arrival_time else "", item.attendance_comment or "", bool(item.photo_data)] for item in recruits],
    )
    _append_sheet(
        workbook,
        "Evaluators",
        ["Evaluator UUID", "Name", "Role", "Active", "Present"],
        [[item.id, item.name, item.role, item.active, item.present] for item in evaluators],
    )
    room_plan = latest_room_plan(db, journey.id, "published")
    room_rows: list[list] = []
    if room_plan:
        room_rows.extend([
            ["Recruit", member.room_number, recruit_by_id.get(member.recruit_id).name if recruit_by_id.get(member.recruit_id) else "", member.recruit_id, ""]
            for member in db.scalars(select(RoomPlanRecruit).where(RoomPlanRecruit.plan_id == room_plan.id))
        ])
        room_rows.extend([
            ["Evaluator", member.room_number, evaluator_by_id.get(member.evaluator_id).name if evaluator_by_id.get(member.evaluator_id) else "", member.evaluator_id, member.mandatory]
            for member in db.scalars(select(RoomPlanEvaluator).where(RoomPlanEvaluator.plan_id == room_plan.id))
        ])
    _append_sheet(workbook, "Rooms", ["Type", "Room", "Name", "UUID", "Mandatory"], room_rows)

    rounds = list(
        db.scalars(
            select(AssignmentRound)
            .where(AssignmentRound.journey_id == journey.id)
            .order_by(AssignmentRound.activity_code, AssignmentRound.version)
        )
    )
    round_by_id = {item.id: item for item in rounds}
    assignments = list(db.scalars(select(Assignment).where(Assignment.round_id.in_(list(round_by_id))))) if rounds else []
    _append_sheet(
        workbook,
        "Assignments",
        ["Activity", "Round version", "Round status", "Evaluator", "Evaluator role", "Recruit", "Room", "Slot", "Repeated", "Reason", "Assignment UUID"],
        [[round_by_id[item.round_id].activity_code, round_by_id[item.round_id].version, round_by_id[item.round_id].status, evaluator_by_id.get(item.evaluator_id).name if evaluator_by_id.get(item.evaluator_id) else "", evaluator_by_id.get(item.evaluator_id).role if evaluator_by_id.get(item.evaluator_id) else "", recruit_by_id.get(item.recruit_id).name if recruit_by_id.get(item.recruit_id) else "", item.room_number, item.slot, item.repeated_pair, item.repeat_reason or "", item.id] for item in assignments],
    )
    submissions = list(db.scalars(select(EvaluationSubmission).where(EvaluationSubmission.journey_id == journey.id)))
    _append_sheet(
        workbook,
        "Evaluations",
        ["Activity", "Evaluator", "Recruit", "Status", "Score /5", "Responses JSON", "Raw JSON", "Comments", "Version", "Submitted UTC", "Submission UUID"],
        [[item.activity_code, evaluator_by_id.get(item.evaluator_id).name if evaluator_by_id.get(item.evaluator_id) else "", recruit_by_id.get(item.recruit_id).name if recruit_by_id.get(item.recruit_id) else "", item.status, float(item.score), item.responses_json, item.raw_payload_json, item.comments, item.version, item.submitted_at.isoformat() if item.submitted_at else "", item.id] for item in submissions],
    )
    admin_evaluations = list(db.scalars(select(AdminEvaluation).where(AdminEvaluation.journey_id == journey.id)))
    _append_sheet(
        workbook,
        "Admin Evaluations",
        ["Activity", "Recruit", "Official score /5", "Responses JSON", "Raw JSON", "Comments", "Updated by", "Version", "Updated UTC"],
        [[item.activity_code, recruit_by_id.get(item.recruit_id).name if recruit_by_id.get(item.recruit_id) else "", float(item.score), item.responses_json, item.raw_payload_json, item.comments, item.updated_by, item.version, item.updated_at.isoformat()] for item in admin_evaluations],
    )
    results = result_snapshot(db, journey)
    _append_sheet(
        workbook,
        "Results",
        [
            "Rank",
            "Recruit",
            "Overall /20",
            "Color",
            "Missing",
            *[DIMENSION_NAMES[code] + " /1" for code in DIMENSION_ORDER],
            "General /1",
            *[RUBRICS[code].name + " /5" for code in ACTIVITY_ORDER],
        ],
        [
            [
                row["overallRank"],
                row["name"],
                row["overallScore"],
                row["color"],
                row["missingCount"],
                *[row["dimensions"][code]["score"] for code in DIMENSION_ORDER],
                row["generalAverage"],
                *[row["activities"][code]["score"] for code in ACTIVITY_ORDER],
            ]
            for row in results["rows"]
        ],
    )
    assessments = list(db.scalars(select(GeneralAssessment).where(GeneralAssessment.recruit_id.in_(list(recruit_by_id))))) if recruits else []
    _append_sheet(
        workbook,
        "General assessments",
        ["Recruit", "Punctuality /1", "Respect /1", "Seriousness /1", "Comment", "Notes", "Version"],
        [[recruit_by_id.get(item.recruit_id).name if recruit_by_id.get(item.recruit_id) else "", float(item.punctuality) if item.punctuality is not None else "", float(item.respect) if item.respect is not None else "", float(item.seriousness) if item.seriousness is not None else "", item.comment, item.notes, item.version] for item in assessments],
    )
    events = list(db.scalars(select(AuditEvent).where(AuditEvent.journey_id == journey.id).order_by(AuditEvent.created_at)))
    _append_sheet(
        workbook,
        "Audit",
        ["Created UTC", "Actor type", "Actor", "Action", "Entity type", "Entity UUID", "Reason", "Before JSON", "After JSON"],
        [[item.created_at.isoformat(), item.actor_type, item.actor_name, item.action, item.entity_type, item.entity_id or "", item.reason, item.before_json, item.after_json] for item in events],
    )
    return workbook


@router.get("/journeys/{journey_id}/export.xlsx")
def export_xlsx(
    journey_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    journey = get_journey_or_404(db, journey_id)
    workbook = build_report_workbook(db, journey, full=True)
    output = io.BytesIO()
    workbook.save(output)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "-", journey.name).strip("-") or "journee"
    return StreamingResponse(
        io.BytesIO(output.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}-full-report.xlsx"',
            "Cache-Control": "no-store",
        },
    )


@router.get("/journeys/{journey_id}/results.xlsx")
def export_results_xlsx(
    journey_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    journey = get_journey_or_404(db, journey_id)
    workbook = build_report_workbook(db, journey, full=False)
    output = io.BytesIO()
    workbook.save(output)
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "-", journey.name).strip("-") or "journee"
    return StreamingResponse(
        io.BytesIO(output.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}-results-and-profiles.xlsx"',
            "Cache-Control": "no-store",
        },
    )


@router.get("/journeys/{journey_id}/results.csv")
def export_results_csv(
    journey_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    journey = get_journey_or_404(db, journey_id)
    results = result_snapshot(db, journey)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Rank",
        "Recruit",
        "Overall /20",
        "Color",
        "Missing",
        *[DIMENSION_NAMES[code] + " /1" for code in DIMENSION_ORDER],
        "General /1",
        *[RUBRICS[code].name + " /5" for code in ACTIVITY_ORDER],
    ])
    for row in results["rows"]:
        writer.writerow([
            row["overallRank"],
            row["name"],
            row["overallScore"],
            row["color"],
            row["missingCount"],
            *[row["dimensions"][code]["score"] for code in DIMENSION_ORDER],
            row["generalAverage"],
            *[row["activities"][code]["score"] for code in ACTIVITY_ORDER],
        ])
    return Response(
        content=output.getvalue().encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="journee-results.csv"'},
    )


@router.get("/journeys/{journey_id}/photos.zip")
def export_photos_zip(
    journey_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    get_journey_or_404(db, journey_id)
    recruits = list(db.scalars(select(Recruit).where(Recruit.journey_id == journey_id)))
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for recruit in recruits:
            if not recruit.photo_data:
                continue
            label = recruit.name
            filename = re.sub(r"[^A-Za-z0-9_-]+", "-", label).strip("-") or recruit.id
            archive.writestr(f"{filename}.webp", recruit.photo_data)
    return StreamingResponse(
        io.BytesIO(output.getvalue()),
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="journee-photos.zip"'},
    )


@router.get("/rubric")
def rubric(context: AdminContext = Depends(require_admin)):
    del context
    return [public_rubric(code) for code in ACTIVITY_ORDER]


@router.get("/journeys")
def list_journeys(
    include_archived: bool = False,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    query = select(Journey).order_by(Journey.updated_at.desc())
    if not include_archived:
        query = query.where(Journey.status != "archived")
    return [serialize_journey(db, journey) for journey in db.scalars(query)]


@router.post("/journeys")
def add_journey(
    payload: JourneyCreateRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = create_journey(db, payload.name, payload.event_date, 1, context.actor_name)
    _commit(db)
    return serialize_journey(db, journey, include_token=True)


@router.get("/journeys/{journey_id}")
def journey_detail(
    journey_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    journey = get_journey_or_404(db, journey_id)
    mandatory_rooms = {
        item.evaluator_id: item.room_number
        for item in db.scalars(
            select(MandatoryRoomEvaluator).where(MandatoryRoomEvaluator.journey_id == journey.id)
        )
    }
    evaluator_payload = []
    for item in db.scalars(
        select(Evaluator).where(Evaluator.journey_id == journey.id)
        .order_by((Evaluator.role == "dossard"), func.lower(Evaluator.name))
    ):
        value = serialize_evaluator(item)
        value["mandatoryRoom"] = mandatory_rooms.get(item.id)
        evaluator_payload.append(value)
    return {
        **serialize_journey(db, journey, include_token=True),
        "activities": _activity_states(db, journey.id),
        "recruits": [
            serialize_recruit(item)
            for item in db.scalars(select(Recruit).where(Recruit.journey_id == journey.id).order_by(Recruit.name))
        ],
        "evaluators": evaluator_payload,
    }


@router.patch("/journeys/{journey_id}")
def update_journey(
    journey_id: str,
    payload: JourneyUpdateRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    if payload.base_version is not None and payload.base_version != journey.version:
        raise HTTPException(status_code=409, detail="This Journee was changed by another admin. Reload before saving.")
    before = serialize_journey(db, journey, include_token=True)
    if payload.name is not None:
        journey.name = " ".join(payload.name.split())
    if payload.event_date is not None:
        journey.event_date = payload.event_date
    if payload.room_count is not None:
        published = latest_room_plan(db, journey.id, "published")
        if published and payload.room_count != journey.room_count:
            raise HTTPException(status_code=409, detail="Room count cannot change after a room plan is published.")
        journey.room_count = payload.room_count
    if payload.status is not None:
        if payload.status == "active":
            other_active = db.scalar(select(Journey).where(Journey.status == "active", Journey.id != journey.id))
            if other_active:
                raise HTTPException(status_code=409, detail=f"{other_active.name} is already active. Complete or archive it first.")
        journey.status = payload.status
        journey.archived_at = utcnow() if payload.status == "archived" else None
    journey.version += 1
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="journey.updated",
        entity_type="journey",
        entity_id=journey.id,
        before=before,
        after={"name": journey.name, "eventDate": journey.event_date, "roomCount": journey.room_count, "status": journey.status},
    )
    _commit(db)
    return serialize_journey(db, journey, include_token=True)


@router.post("/journeys/{journey_id}/duplicate")
def duplicate_journey(
    journey_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    source = get_journey_or_404(db, journey_id)
    clone = create_journey(db, f"{source.name} Copy", source.event_date, source.room_count, context.actor_name)
    evaluator_id_map: dict[str, str] = {}
    for item in db.scalars(select(Recruit).where(Recruit.journey_id == source.id, Recruit.active.is_(True))):
        db.add(Recruit(journey_id=clone.id, name=item.name, phone_number=item.phone_number,
                       date_of_birth=item.date_of_birth, attendance_comment=item.attendance_comment,
                       photo_data=item.photo_data, photo_type=item.photo_type))
    clone_evaluators = list(db.scalars(select(Evaluator).where(Evaluator.journey_id == clone.id)))
    clone_by_directory = {item.directory_id: item for item in clone_evaluators if item.directory_id}
    clone_by_name = {item.name.casefold(): item for item in clone_evaluators}
    for item in db.scalars(select(Evaluator).where(Evaluator.journey_id == source.id, Evaluator.active.is_(True))):
        copied = clone_by_directory.get(item.directory_id) or clone_by_name.get(item.name.casefold())
        if copied:
            copied.role = item.role
        else:
            copied = Evaluator(journey_id=clone.id, directory_id=item.directory_id, name=item.name, role=item.role)
            db.add(copied)
            db.flush()
        evaluator_id_map[item.id] = copied.id
    for item in db.scalars(select(MandatoryRoomEvaluator).where(MandatoryRoomEvaluator.journey_id == source.id)):
        if item.evaluator_id in evaluator_id_map:
            db.add(MandatoryRoomEvaluator(journey_id=clone.id, evaluator_id=evaluator_id_map[item.evaluator_id], room_number=item.room_number))
    _commit(db)
    return serialize_journey(db, clone, include_token=True)


@router.post("/journeys/{journey_id}/rotate-link")
def rotate_public_link(
    journey_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    journey.public_token = secrets.token_urlsafe(32)
    journey.version += 1
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="journey.public_link_rotated",
        entity_type="journey",
        entity_id=journey.id,
    )
    _commit(db)
    return {"publicToken": journey.public_token, "evaluatorPath": f"/j/{journey.public_token}"}


@router.post("/journeys/{journey_id}/rotate-recruit-attendance-link")
def rotate_recruit_attendance_link(
    journey_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    access = db.get(RecruitAttendanceAccess, journey.id)
    if not access:
        access = RecruitAttendanceAccess(journey_id=journey.id, token=secrets.token_urlsafe(32))
        db.add(access)
    else:
        access.token = secrets.token_urlsafe(32)
        access.rotated_at = utcnow()
    db.execute(
        delete(RecruitAttendanceSession).where(RecruitAttendanceSession.journey_id == journey.id)
    )
    journey.version += 1
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="journey.recruit_attendance_link_rotated",
        entity_type="journey",
        entity_id=journey.id,
    )
    _commit(db)
    return {"recruitAttendancePath": f"/recruit-attendance/{access.token}"}


@router.get("/journeys/{journey_id}/dashboard")
def dashboard(
    journey_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    journey = get_journey_or_404(db, journey_id)
    evaluators = list(
        db.scalars(select(Evaluator).where(Evaluator.journey_id == journey.id, Evaluator.active.is_(True), Evaluator.present.is_(True)))
    )
    room_plan = latest_room_plan(db, journey.id, "published")
    results = result_snapshot(db, journey)
    warnings: list[str] = loads(room_plan.warnings_json, []) if room_plan else []
    states = list(db.scalars(select(ActivityState).where(ActivityState.journey_id == journey.id)))
    for state in states:
        if state.assignment_round_id:
            round_record = db.get(AssignmentRound, state.assignment_round_id)
            if round_record:
                warnings.extend(
                    f"{RUBRICS[state.code].name}: {message}"
                    for message in loads(round_record.warnings_json, [])
                    if message not in {
                        "Shared Skills & Simulation assignment.",
                        "Simulation reuses the published Skills assignment.",
                    }
                )
    active_monitoring = monitoring_snapshot(db, journey, journey.current_activity) if journey.current_activity else None
    return {
        "journey": serialize_journey(db, journey, include_token=True),
        "activities": _activity_states(db, journey.id),
        "overallEvaluators": sum(1 for item in evaluators if item.role == "overall"),
        "dossardEvaluators": sum(1 for item in evaluators if item.role == "dossard"),
        "roomPlan": room_plan_payload(db, room_plan) if room_plan else None,
        "dimensionAverages": results["dimensionAverages"],
        "dimensionNames": results["dimensionNames"],
        "formula": results["formula"],
        "activityAverages": results["activityAverages"],
        "ranking": results["rows"][:10],
        "warnings": warnings,
        "activeMonitoring": active_monitoring,
    }


@router.post("/journeys/{journey_id}/recruits")
def add_recruit(
    journey_id: str,
    payload: RecruitCreateRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    recruit = Recruit(
        journey_id=journey.id,
        name=" ".join(payload.name.split()),
        phone_number=(payload.phone_number or "").strip() or None,
        date_of_birth=payload.date_of_birth,
    )
    db.add(recruit)
    db.flush()
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name, action="recruit.added", entity_type="recruit", entity_id=recruit.id, after=serialize_recruit(recruit))
    _commit(db)
    return serialize_recruit(recruit)


@router.post("/journeys/{journey_id}/evaluators")
def add_evaluator(
    journey_id: str,
    payload: EvaluatorCreateRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    clean_name = " ".join(payload.name.split())
    existing = db.scalar(
        select(Evaluator).where(
            Evaluator.journey_id == journey.id,
            func.lower(Evaluator.name) == clean_name.lower(),
        )
    )
    if existing:
        raise HTTPException(status_code=409, detail="This evaluator is already in the Journee directory.")
    directory = db.scalar(select(EvaluatorDirectory).where(func.lower(EvaluatorDirectory.name) == clean_name.lower()))
    if payload.add_to_directory:
        if not directory:
            directory = EvaluatorDirectory(name=clean_name, default_role=payload.role)
            db.add(directory)
            db.flush()
        elif not directory.active:
            directory.name = clean_name
            directory.default_role = payload.role
            directory.active = True
    account = db.scalar(select(UserAccount).where(func.lower(UserAccount.username) == clean_name.lower()))
    if payload.add_to_directory and not account:
        if not payload.password:
            raise HTTPException(status_code=422, detail="A password is required for a new evaluator account.")
        account = UserAccount(
            username=clean_name,
            password_hash=hash_password(payload.password),
            directory_id=directory.id,
            evaluator_role=payload.role,
            must_change_password=True,
        )
        db.add(account)
    elif account:
        account.active = True
        account.evaluator_role = payload.role
    evaluator = Evaluator(journey_id=journey.id, directory_id=directory.id if directory else None, name=clean_name, role=payload.role)
    db.add(evaluator)
    db.flush()
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name, action="evaluator.added", entity_type="evaluator", entity_id=evaluator.id, after=serialize_evaluator(evaluator))
    _commit(db)
    return serialize_evaluator(evaluator)


@router.put("/journeys/{journey_id}/attendance/recruits")
def save_recruit_attendance(
    journey_id: str,
    payload: RecruitAttendanceRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    changed = []
    for item in payload.items:
        recruit = get_recruit_or_404(db, journey.id, item.id)
        if item.base_version is not None and item.base_version != recruit.version:
            raise HTTPException(status_code=409, detail=f"{recruit.name} was changed by another admin. Reload attendance.")
        before = serialize_recruit(recruit)
        recruit.present = item.present
        recruit.arrival_time = item.arrival_time if item.present else None
        recruit.phone_number = (item.phone_number or "").strip() or None
        if "date_of_birth" in item.model_fields_set:
            recruit.date_of_birth = item.date_of_birth
        recruit.attendance_comment = item.attendance_comment.strip()
        recruit.active = item.active
        recruit.version += 1
        changed.append({"before": before, "after": serialize_recruit(recruit)})
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name, action="attendance.recruits_saved", entity_type="attendance", after=changed)
    _commit(db)
    return {"ok": True, "items": [serialize_recruit(get_recruit_or_404(db, journey.id, item.id)) for item in payload.items]}


@router.patch("/journeys/{journey_id}/recruits/{recruit_id}/attendance")
def auto_save_recruit_attendance(
    journey_id: str,
    recruit_id: str,
    payload: RecruitAttendancePatchRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    changes = payload.model_dump(exclude={"base_version"}, exclude_unset=True)
    before, recruit = update_recruit_attendance_fields(
        db, journey.id, recruit_id, payload.base_version, changes
    )
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="attendance.recruit_auto_saved",
        entity_type="recruit",
        entity_id=recruit.id,
        before=before,
        after=serialize_recruit(recruit),
    )
    _commit(db)
    return serialize_recruit(recruit)


@router.put("/journeys/{journey_id}/attendance/evaluators")
def save_evaluator_attendance(
    journey_id: str,
    payload: EvaluatorAttendanceRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    changed = []
    for item in payload.items:
        evaluator = get_evaluator_or_404(db, journey.id, item.id)
        if item.base_version is not None and item.base_version != evaluator.version:
            raise HTTPException(status_code=409, detail=f"{evaluator.name} was changed by another admin. Reload attendance.")
        before = serialize_evaluator(evaluator)
        evaluator.present = item.present
        evaluator.role = item.role
        evaluator.active = item.active
        evaluator.version += 1
        changed.append({"before": before, "after": serialize_evaluator(evaluator)})
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name, action="attendance.evaluators_saved", entity_type="attendance", after=changed)
    _commit(db)
    return {"ok": True, "items": [serialize_evaluator(get_evaluator_or_404(db, journey.id, item.id)) for item in payload.items]}


@router.delete("/journeys/{journey_id}/recruits/{recruit_id}")
def remove_recruit(
    journey_id: str,
    recruit_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    recruit = get_recruit_or_404(db, journey.id, recruit_id)
    referenced = bool(
        db.scalar(select(Assignment.id).where(Assignment.recruit_id == recruit.id).limit(1))
        or db.scalar(select(RoomPlanRecruit.id).where(RoomPlanRecruit.recruit_id == recruit.id).limit(1))
        or db.get(GeneralAssessment, recruit.id)
    )
    disposition = "deactivated" if referenced else "deleted"
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name,
          action=f"recruit.{disposition}", entity_type="recruit", entity_id=recruit.id,
          before=serialize_recruit(recruit), reason="Removed from attendance by admin.")
    if referenced:
        recruit.active = False
        recruit.present = False
        recruit.version += 1
    else:
        db.delete(recruit)
    _commit(db)
    return {"ok": True, "disposition": disposition}


@router.delete("/journeys/{journey_id}/evaluators/{evaluator_id}")
def remove_evaluator(
    journey_id: str,
    evaluator_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    evaluator = get_evaluator_or_404(db, journey.id, evaluator_id)
    referenced = bool(
        db.scalar(select(Assignment.id).where(Assignment.evaluator_id == evaluator.id).limit(1))
        or db.scalar(select(RoomPlanEvaluator.id).where(RoomPlanEvaluator.evaluator_id == evaluator.id).limit(1))
    )
    db.execute(delete(MandatoryRoomEvaluator).where(MandatoryRoomEvaluator.journey_id == journey.id,
                                                     MandatoryRoomEvaluator.evaluator_id == evaluator.id))
    disposition = "deactivated" if referenced else "deleted"
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name,
          action=f"evaluator.{disposition}", entity_type="evaluator", entity_id=evaluator.id,
          before=serialize_evaluator(evaluator), reason="Removed from attendance by admin.")
    if referenced:
        evaluator.active = False
        evaluator.present = False
        evaluator.version += 1
    else:
        db.delete(evaluator)
    _commit(db)
    return {"ok": True, "disposition": disposition}


@router.post("/journeys/{journey_id}/recruits/{recruit_id}/photo")
def upload_recruit_photo(
    journey_id: str,
    recruit_id: str,
    request: Request,
    photo: UploadFile = File(...),
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    recruit = get_recruit_or_404(db, journey_id, recruit_id)
    data, media_type = process_photo(photo.file.read())
    recruit.photo_data = data
    recruit.photo_type = media_type
    recruit.photo_updated_at = utcnow()
    recruit.version += 1
    audit(db, journey_id=journey_id, actor_type="admin", actor_name=context.actor_name, action="recruit.photo_updated", entity_type="recruit", entity_id=recruit.id)
    _commit(db)
    return {"ok": True, "hasPhoto": True, "version": recruit.version}


@router.get("/journeys/{journey_id}/recruits/{recruit_id}/photo")
def admin_recruit_photo(
    journey_id: str,
    recruit_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    recruit = get_recruit_or_404(db, journey_id, recruit_id)
    if not recruit.photo_data:
        raise HTTPException(status_code=404, detail="No photo available.")
    return Response(content=recruit.photo_data, media_type=recruit.photo_type or "image/webp", headers={"Cache-Control": "private, max-age=300"})


@router.put("/journeys/{journey_id}/mandatory-rooms")
def save_mandatory_rooms(
    journey_id: str,
    payload: MandatoryRoomRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    if latest_room_plan(db, journey.id, "published"):
        raise HTTPException(status_code=409, detail="Mandatory rooms cannot change after a room plan is published.")
    seen: set[str] = set()
    for item in payload.items:
        get_evaluator_or_404(db, journey.id, item.evaluator_id)
        if item.evaluator_id in seen:
            raise HTTPException(status_code=422, detail="An evaluator can be mandatory in only one room.")
        if item.room_number > journey.room_count:
            raise HTTPException(status_code=422, detail="Mandatory room exceeds the Journee room count.")
        seen.add(item.evaluator_id)
    db.execute(delete(MandatoryRoomEvaluator).where(MandatoryRoomEvaluator.journey_id == journey.id))
    for item in payload.items:
        db.add(MandatoryRoomEvaluator(journey_id=journey.id, evaluator_id=item.evaluator_id, room_number=item.room_number))
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name, action="rooms.mandatory_saved", entity_type="mandatory_room", after=[item.model_dump() for item in payload.items])
    _commit(db)
    return {"ok": True}


@router.get("/journeys/{journey_id}/rooms")
def get_rooms(
    journey_id: str,
    status: str = Query(default="published", pattern="^(preview|published)$"),
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    get_journey_or_404(db, journey_id)
    plan = latest_room_plan(db, journey_id, status)
    return room_plan_payload(db, plan) if plan else None


@router.put("/journeys/{journey_id}/rooms/count")
def change_room_count(
    journey_id: str,
    payload: RoomCountRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    if payload.room_count == journey.room_count:
        return {"ok": True, "roomCount": journey.room_count}
    open_room_activity = db.scalar(select(ActivityState).where(
        ActivityState.journey_id == journey.id,
        ActivityState.code.in_(("escape_room", "negotiation")),
        ActivityState.status == "open",
    ))
    if open_room_activity:
        raise HTTPException(status_code=409, detail="Close Escape Room and Negotiation before changing the room count.")
    published = latest_room_plan(db, journey.id, "published")
    if published and not payload.reason.strip():
        raise HTTPException(status_code=422, detail="A reason is required because a room plan is already published.")
    present_recruits = db.scalar(select(func.count()).select_from(Recruit).where(
        Recruit.journey_id == journey.id, Recruit.active.is_(True), Recruit.present.is_(True))) or 0
    if present_recruits and payload.room_count > present_recruits:
        raise HTTPException(status_code=422, detail="Room count cannot exceed the number of present recruits.")
    invalid_mandatory = db.scalar(select(MandatoryRoomEvaluator).where(
        MandatoryRoomEvaluator.journey_id == journey.id,
        MandatoryRoomEvaluator.room_number > payload.room_count))
    if invalid_mandatory:
        raise HTTPException(status_code=422, detail="Move mandatory evaluators out of rooms above the new room count first.")
    before = journey.room_count
    db.execute(update(RoomPlan).where(RoomPlan.journey_id == journey.id,
                                      RoomPlan.status.in_(("preview", "published"))).values(status="superseded"))
    journey.room_count = payload.room_count
    journey.version += 1
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name,
          action="rooms.count_changed", entity_type="journey", entity_id=journey.id,
          before={"roomCount": before}, after={"roomCount": journey.room_count}, reason=payload.reason)
    _commit(db)
    return {"ok": True, "roomCount": journey.room_count}


@router.post("/journeys/{journey_id}/rooms/preview")
def preview_rooms(
    journey_id: str,
    payload: PreviewRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    seed = payload.seed or secrets.token_hex(8)
    plan = create_room_preview(db, journey, context.actor_name, seed)
    _commit(db)
    return room_plan_payload(db, plan)


@router.put("/journeys/{journey_id}/rooms/{plan_id}")
def edit_room_preview(
    journey_id: str,
    plan_id: str,
    payload: RoomPlanEditRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    plan = db.get(RoomPlan, plan_id)
    if not plan or plan.journey_id != journey.id or plan.status != "preview":
        raise HTTPException(status_code=409, detail="Editable room preview not found.")
    expected_recruits = set(
        db.scalars(select(Recruit.id).where(Recruit.journey_id == journey.id, Recruit.active.is_(True), Recruit.present.is_(True)))
    )
    expected_evaluators = set(
        db.scalars(select(Evaluator.id).where(Evaluator.journey_id == journey.id, Evaluator.active.is_(True), Evaluator.present.is_(True)))
    )
    if set(payload.recruit_rooms) != expected_recruits or set(payload.evaluator_rooms) != expected_evaluators:
        raise HTTPException(status_code=422, detail="The room preview must include every present recruit and evaluator exactly once.")
    if any(room < 1 or room > journey.room_count for room in [*payload.recruit_rooms.values(), *payload.evaluator_rooms.values()]):
        raise HTTPException(status_code=422, detail="One or more room numbers are invalid.")
    db.execute(delete(RoomPlanRecruit).where(RoomPlanRecruit.plan_id == plan.id))
    db.execute(delete(RoomPlanEvaluator).where(RoomPlanEvaluator.plan_id == plan.id))
    mandatory_rooms = {item.evaluator_id: item.room_number for item in db.scalars(
        select(MandatoryRoomEvaluator).where(MandatoryRoomEvaluator.journey_id == journey.id))}
    for evaluator_id, mandatory_room in mandatory_rooms.items():
        if payload.evaluator_rooms.get(evaluator_id) != mandatory_room:
            evaluator = get_evaluator_or_404(db, journey.id, evaluator_id)
            raise HTTPException(status_code=422, detail=f"{evaluator.name} is mandatory in Room {mandatory_room} and cannot be moved.")
    for recruit_id, room in payload.recruit_rooms.items():
        db.add(RoomPlanRecruit(plan_id=plan.id, recruit_id=recruit_id, room_number=room))
    for evaluator_id, room in payload.evaluator_rooms.items():
        db.add(RoomPlanEvaluator(plan_id=plan.id, evaluator_id=evaluator_id, room_number=room, mandatory=evaluator_id in mandatory_rooms))
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name, action="room_plan.manually_edited", entity_type="room_plan", entity_id=plan.id)
    _commit(db)
    return room_plan_payload(db, plan)


@router.post("/journeys/{journey_id}/rooms/{plan_id}/publish")
def confirm_room_plan(
    journey_id: str,
    plan_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    plan = db.get(RoomPlan, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Room preview not found.")
    publish_room_plan(db, journey, plan, context.actor_name)
    _commit(db)
    return room_plan_payload(db, plan)


@router.post("/journeys/{journey_id}/rooms/{plan_id}/publish-override")
def override_room_plan(
    journey_id: str,
    plan_id: str,
    payload: ActivityActionRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    if not payload.reason.strip():
        raise HTTPException(status_code=422, detail="A reason is required for a room override.")
    journey = get_journey_or_404(db, journey_id)
    plan = db.get(RoomPlan, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Room preview not found.")
    publish_room_plan(db, journey, plan, context.actor_name, override_reason=payload.reason)
    _commit(db)
    return room_plan_payload(db, plan)


@router.get("/journeys/{journey_id}/assignments/{activity_code}")
def get_assignments(
    journey_id: str,
    activity_code: str,
    status: str = Query(default="published", pattern="^(preview|published)$"),
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    get_journey_or_404(db, journey_id)
    round_record = db.scalar(
        select(AssignmentRound)
        .where(
            AssignmentRound.journey_id == journey_id,
            AssignmentRound.activity_code == activity_code,
            AssignmentRound.status == status,
        )
        .order_by(AssignmentRound.version.desc())
    )
    return assignment_round_payload(db, round_record) if round_record else None


@router.post("/journeys/{journey_id}/assignments/{activity_code}/preview")
def preview_assignments(
    journey_id: str,
    activity_code: str,
    payload: PreviewRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    seed = payload.seed or secrets.token_hex(8)
    round_record = create_assignment_preview(db, journey, activity_code, context.actor_name, seed)
    _commit(db)
    return assignment_round_payload(db, round_record)


@router.put("/journeys/{journey_id}/assignments/{round_id}")
def edit_assignment_preview(
    journey_id: str,
    round_id: str,
    payload: AssignmentEditRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    round_record = db.get(AssignmentRound, round_id)
    if not round_record or round_record.journey_id != journey.id or round_record.status != "preview":
        raise HTTPException(status_code=409, detail="Editable assignment preview not found.")
    if round_record.activity_code == "simulation":
        raise HTTPException(status_code=409, detail="Simulation must reuse the Skills assignments exactly and cannot be edited independently.")
    pair_keys: set[tuple[str, str]] = set()
    recruit_counts: dict[str, int] = Counter()
    evaluator_counts: dict[str, int] = Counter()
    past_pairs, _secondary_counts = past_pair_data(db, journey.id, round_record.activity_code)
    for item in payload.items:
        evaluator = get_evaluator_or_404(db, journey.id, item.evaluator_id)
        recruit = get_recruit_or_404(db, journey.id, item.recruit_id)
        if not evaluator.present or not recruit.present or not evaluator.active or not recruit.active:
            raise HTTPException(status_code=422, detail="Assignments may contain only confirmed-present active people.")
        pair = (item.evaluator_id, item.recruit_id)
        if pair in pair_keys:
            raise HTTPException(status_code=422, detail="Duplicate evaluator–recruit pair.")
        pair_keys.add(pair)
        recruit_counts[item.recruit_id] += 1
        evaluator_counts[item.evaluator_id] += 1
        if recruit_counts[item.recruit_id] > 2:
            raise HTTPException(status_code=422, detail="A recruit cannot have more than two evaluators.")
        if pair in past_pairs and not (item.override_reason or "").strip():
            raise HTTPException(status_code=422, detail="A repeated evaluator-recruit pair requires an override reason.")
    present_recruits = set(
        db.scalars(
            select(Recruit.id).where(
                Recruit.journey_id == journey.id,
                Recruit.active.is_(True),
                Recruit.present.is_(True),
            )
        )
    )
    if set(recruit_counts) != present_recruits or any(recruit_counts[item] < 1 for item in present_recruits):
        raise HTTPException(status_code=422, detail="Every confirmed-present recruit must retain a primary assignment.")
    for recruit_id in present_recruits:
        slots = [item.slot for item in payload.items if item.recruit_id == recruit_id]
        if slots.count(1) != 1 or slots.count(2) > 1:
            raise HTTPException(status_code=422, detail="Each recruit needs exactly one primary slot and at most one secondary slot.")
    present_evaluator_count = db.scalar(
        select(func.count()).select_from(Evaluator).where(
            Evaluator.journey_id == journey.id,
            Evaluator.active.is_(True),
            Evaluator.present.is_(True),
        )
    ) or 0
    if round_record.activity_code not in ROOM_ACTIVITIES:
        if present_evaluator_count >= len(present_recruits) and any(value > 1 for value in evaluator_counts.values()):
            raise HTTPException(status_code=422, detail="With no evaluator shortage, an evaluator may have only one recruit.")
        if 0 < present_evaluator_count < len(present_recruits):
            capacity = (len(present_recruits) + present_evaluator_count - 1) // present_evaluator_count
            if any(value > capacity for value in evaluator_counts.values()):
                raise HTTPException(status_code=422, detail="A manual edit exceeds the balanced shortage capacity.")
    if round_record.activity_code in ROOM_ACTIVITIES:
        room_plan = latest_room_plan(db, journey.id, "published")
        if not room_plan:
            raise HTTPException(status_code=409, detail="Published room plan not found.")
        recruit_rooms = {item.recruit_id: item.room_number for item in db.scalars(select(RoomPlanRecruit).where(RoomPlanRecruit.plan_id == room_plan.id))}
        evaluator_rooms = {item.evaluator_id: item.room_number for item in db.scalars(select(RoomPlanEvaluator).where(RoomPlanEvaluator.plan_id == room_plan.id))}
        room_recruit_counts = Counter(recruit_rooms.values())
        room_evaluator_counts = Counter(evaluator_rooms.values())
        for item in payload.items:
            if recruit_rooms.get(item.recruit_id) != evaluator_rooms.get(item.evaluator_id):
                raise HTTPException(status_code=422, detail="Room activities can pair only people in the same room.")
        for evaluator_id, load in evaluator_counts.items():
            room = evaluator_rooms[evaluator_id]
            capacity = max(1, (room_recruit_counts[room] + max(room_evaluator_counts[room], 1) - 1) // max(room_evaluator_counts[room], 1))
            if load > capacity:
                raise HTTPException(status_code=422, detail="A manual edit exceeds the balanced evaluator capacity for its room.")
    db.execute(delete(Assignment).where(Assignment.round_id == round_record.id))
    for item in payload.items:
        room_number = recruit_rooms[item.recruit_id] if round_record.activity_code in ROOM_ACTIVITIES else None
        repeated = (item.evaluator_id, item.recruit_id) in past_pairs
        db.add(Assignment(round_id=round_record.id, evaluator_id=item.evaluator_id, recruit_id=item.recruit_id, room_number=room_number, slot=item.slot, repeated_pair=repeated, repeat_reason=item.override_reason if repeated else None))
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name, action="assignments.manually_edited", entity_type="assignment_round", entity_id=round_record.id)
    _commit(db)
    return assignment_round_payload(db, round_record)


@router.post("/journeys/{journey_id}/assignments/{round_id}/publish")
def confirm_assignments(
    journey_id: str,
    round_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    round_record = db.get(AssignmentRound, round_id)
    if not round_record:
        raise HTTPException(status_code=404, detail="Assignment preview not found.")
    publish_assignment_round(db, journey, round_record, context.actor_name)
    _commit(db)
    return assignment_round_payload(db, round_record)


@router.post("/journeys/{journey_id}/activities/skills-simulation/{action}")
def paired_skills_simulation_action(
    journey_id: str,
    action: str,
    payload: ActivityActionRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    if action not in {"open", "close", "reopen"}:
        raise HTTPException(status_code=404, detail="Unknown paired activity action.")
    journey = get_journey_or_404(db, journey_id)
    states = {item.code: item for item in db.scalars(select(ActivityState).where(
        ActivityState.journey_id == journey.id, ActivityState.code.in_(("skills", "simulation"))))}
    if set(states) != {"skills", "simulation"}:
        raise HTTPException(status_code=404, detail="Skills or Simulation state is missing.")
    if action in {"open", "reopen"}:
        if action == "reopen" and not payload.reason.strip():
            raise HTTPException(status_code=422, detail="A reason is required to reopen both activities.")
        if any(not states[code].assignment_round_id for code in states):
            raise HTTPException(status_code=409, detail="Publish both Skills and Simulation assignments first.")
        negotiation = db.scalar(select(ActivityState).where(ActivityState.journey_id == journey.id,
                                                              ActivityState.code == "negotiation"))
        if not negotiation or negotiation.status != "closed":
            raise HTTPException(status_code=409, detail="Close Negotiation before opening Skills and Simulation.")
        conflicting = db.scalar(select(ActivityState).where(
            ActivityState.journey_id == journey.id, ActivityState.status == "open",
            ActivityState.code.not_in(("skills", "simulation"))))
        if conflicting:
            raise HTTPException(status_code=409, detail=f"Close {RUBRICS[conflicting.code].name} first.")
        other_active = db.scalar(select(Journey).where(Journey.status == "active", Journey.id != journey.id))
        if other_active:
            raise HTTPException(status_code=409, detail=f"{other_active.name} is already active. Complete or archive it first.")
        for state in states.values():
            state.status, state.opened_at, state.closed_at = "open", utcnow(), None
            state.version += 1
        journey.current_activity, journey.status = "skills", "active"
        db.execute(update(EvaluationSubmission).where(
            EvaluationSubmission.journey_id == journey.id,
            EvaluationSubmission.activity_code.in_(("skills", "simulation")),
            EvaluationSubmission.status == "locked").values(status="submitted"))
    else:
        if not any(state.status == "open" for state in states.values()):
            raise HTTPException(status_code=409, detail="Neither Skills nor Simulation is open.")
        for state in states.values():
            if state.status == "open":
                state.status, state.closed_at = "closed", utcnow()
                state.version += 1
        journey.current_activity = None
        db.execute(update(EvaluationSubmission).where(
            EvaluationSubmission.journey_id == journey.id,
            EvaluationSubmission.activity_code.in_(("skills", "simulation")),
            EvaluationSubmission.status == "submitted").values(status="locked"))
    journey.version += 1
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name,
          action=f"activity.skills_simulation_{action}", entity_type="activity", entity_id=None,
          after={"activities": ["skills", "simulation"], "status": "open" if action != "close" else "closed"},
          reason=payload.reason)
    _commit(db)
    return {"ok": True, "activities": ["skills", "simulation"],
            "status": "open" if action != "close" else "closed"}


@router.post("/journeys/{journey_id}/activities/{activity_code}/{action}")
def activity_action(
    journey_id: str,
    activity_code: str,
    action: str,
    payload: ActivityActionRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    if action not in {"open", "close", "reopen"} or activity_code not in RUBRICS:
        raise HTTPException(status_code=404, detail="Unknown activity action.")
    journey = get_journey_or_404(db, journey_id)
    state = db.scalar(select(ActivityState).where(ActivityState.journey_id == journey.id, ActivityState.code == activity_code))
    if not state:
        raise HTTPException(status_code=404, detail="Activity not found.")
    if action in {"open", "reopen"}:
        other_active = db.scalar(select(Journey).where(Journey.status == "active", Journey.id != journey.id))
        if other_active:
            raise HTTPException(status_code=409, detail=f"{other_active.name} is already active. Complete or archive it first.")
        if action == "reopen" and not payload.reason.strip():
            raise HTTPException(status_code=422, detail="A reason is required to reopen an activity.")
        if not state.assignment_round_id:
            raise HTTPException(status_code=409, detail="Publish assignments before opening this activity.")
        predecessor = {"negotiation": "escape_room", "skills": "negotiation", "simulation": "negotiation"}.get(activity_code)
        if predecessor:
            predecessor_state = db.scalar(
                select(ActivityState).where(
                    ActivityState.journey_id == journey.id,
                    ActivityState.code == predecessor,
                )
            )
            if not predecessor_state or predecessor_state.status != "closed":
                raise HTTPException(status_code=409, detail=f"Close {RUBRICS[predecessor].name} before opening {RUBRICS[activity_code].name}.")
        another = db.scalar(select(ActivityState).where(ActivityState.journey_id == journey.id, ActivityState.status == "open", ActivityState.code != activity_code))
        if another and {another.code, activity_code} != {"skills", "simulation"}:
            raise HTTPException(status_code=409, detail=f"Close {RUBRICS[another.code].name} before opening another activity.")
        state.status = "open"
        state.opened_at = utcnow()
        state.closed_at = None
        journey.current_activity = "skills" if activity_code in {"skills", "simulation"} and another else activity_code
        journey.status = "active"
        db.execute(
            update(EvaluationSubmission)
            .where(EvaluationSubmission.journey_id == journey.id, EvaluationSubmission.activity_code == activity_code, EvaluationSubmission.status == "locked")
            .values(status="submitted")
        )
    else:
        if state.status != "open":
            raise HTTPException(status_code=409, detail="Only an open activity can be closed.")
        state.status = "closed"
        state.closed_at = utcnow()
        other_open = db.scalar(select(ActivityState).where(ActivityState.journey_id == journey.id,
                                                            ActivityState.status == "open",
                                                            ActivityState.code != activity_code))
        journey.current_activity = other_open.code if other_open else None
        db.execute(
            update(EvaluationSubmission)
            .where(EvaluationSubmission.journey_id == journey.id, EvaluationSubmission.activity_code == activity_code, EvaluationSubmission.status == "submitted")
            .values(status="locked")
        )
    state.version += 1
    journey.version += 1
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name, action=f"activity.{action}", entity_type="activity", entity_id=state.id, after={"activity": activity_code, "status": state.status}, reason=payload.reason)
    _commit(db)
    return {"ok": True, "activity": activity_code, "status": state.status}


def _simulated_evaluation_values(activity_code: str, rng: random.Random) -> tuple[dict, dict]:
    rubric = RUBRICS[activity_code]
    if rubric.kind == "sport":
        return {}, {
            "push_ups": rng.randint(8, 40),
            "wall_sit": rng.randint(35, 150),
            "beep_test": rng.randint(240, 720),
            "plank": rng.randint(30, 180),
        }
    return {
        criterion.key: rng.randint(10, 50) / 10
        for criterion in rubric.criteria
    }, {}


@router.post("/journeys/{journey_id}/testing/simulate/{activity_code}")
def simulate_open_activity_evaluations(
    journey_id: str,
    activity_code: str,
    payload: EvaluationSimulationRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    if not settings.allow_test_tools:
        raise HTTPException(status_code=404, detail="Testing tools are not enabled.")
    if activity_code not in RUBRICS:
        raise HTTPException(status_code=404, detail="Unknown activity.")
    journey = get_journey_or_404(db, journey_id)
    request_key = request.headers.get("Idempotency-Key", "").strip()[:120]
    idempotency_scope = f"testing-simulator:{journey.id}:{activity_code}"
    if request_key:
        prior = db.scalar(select(IdempotencyRecord).where(
            IdempotencyRecord.scope == idempotency_scope,
            IdempotencyRecord.request_key == request_key,
        ))
        if prior:
            return loads(prior.response_json, {"ok": True})
    state = db.scalar(
        select(ActivityState)
        .where(ActivityState.journey_id == journey.id, ActivityState.code == activity_code)
        .with_for_update()
    )
    if not state or state.status != "open":
        raise HTTPException(status_code=409, detail="Open the activity before running simulated evaluations.")
    round_record = db.get(AssignmentRound, state.assignment_round_id) if state.assignment_round_id else None
    if not round_record or round_record.status != "published":
        raise HTTPException(status_code=409, detail="A published assignment is required before simulation.")

    assignments = list(
        db.scalars(
            select(Assignment)
            .where(Assignment.round_id == round_record.id)
            .order_by(Assignment.id)
            .with_for_update()
        )
    )
    submitted_ids = set(
        db.scalars(
            select(EvaluationSubmission.assignment_id).where(
                EvaluationSubmission.assignment_id.in_([item.id for item in assignments]),
                EvaluationSubmission.status.in_(("submitted", "locked")),
            )
        )
    ) if assignments else set()
    remaining = [item for item in assignments if item.id not in submitted_ids]
    if not remaining:
        raise HTTPException(status_code=409, detail="Every assigned evaluation is already submitted.")
    requested = payload.count if payload.count is not None else len(remaining)
    if requested > len(remaining):
        raise HTTPException(status_code=422, detail=f"Only {len(remaining)} evaluation(s) remain unsubmitted.")

    rng = random.Random(secrets.randbits(128))
    rng.shuffle(remaining)
    selected = remaining[:requested]
    evaluator_by_id = {
        item.id: item for item in db.scalars(
            select(Evaluator).where(Evaluator.id.in_({assignment.evaluator_id for assignment in selected}))
        )
    }
    recruit_by_id = {
        item.id: item for item in db.scalars(
            select(Recruit).where(Recruit.id.in_({assignment.recruit_id for assignment in selected}))
        )
    }
    completed = []
    for assignment in selected:
        evaluator = evaluator_by_id[assignment.evaluator_id]
        recruit = recruit_by_id[assignment.recruit_id]
        responses, raw = _simulated_evaluation_values(activity_code, rng)
        submission = save_submission(
            db,
            assignment=assignment,
            round_record=round_record,
            state=state,
            responses=responses,
            raw=raw,
            comments="Automated test evaluation generated from the evaluator workflow simulator.",
            actor_type="evaluator",
            actor_name=evaluator.name,
            submit=True,
            reason="Generated by the development-only admin testing tool.",
        )
        completed.append({
            "assignmentId": assignment.id,
            "submissionId": submission.id,
            "evaluatorId": evaluator.id,
            "evaluatorName": evaluator.name,
            "recruitId": recruit.id,
            "recruitName": recruit.name,
            "score": float(submission.score),
        })
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="testing.evaluations_simulated",
        entity_type="activity",
        entity_id=state.id,
        after={
            "activity": activity_code,
            "requested": requested,
            "completed": completed,
            "remainingAfter": len(remaining) - requested,
        },
        reason="Development-only evaluator workflow simulation.",
    )
    result = {
        "ok": True,
        "activityCode": activity_code,
        "completedCount": len(completed),
        "remainingCount": len(remaining) - len(completed),
        "completed": completed,
    }
    if request_key:
        db.add(IdempotencyRecord(
            scope=idempotency_scope,
            request_key=request_key,
            response_json=dumps(result),
        ))
    _commit(db)
    return result


@router.get("/journeys/{journey_id}/monitoring/{activity_code}")
def activity_monitoring(
    journey_id: str,
    activity_code: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    if activity_code not in RUBRICS:
        raise HTTPException(status_code=404, detail="Unknown activity.")
    return monitoring_snapshot(db, get_journey_or_404(db, journey_id), activity_code)


@router.get("/journeys/{journey_id}/results")
def journey_results(
    journey_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    return result_snapshot(db, get_journey_or_404(db, journey_id))


def _submission_payload(db: Session, submission: EvaluationSubmission, evaluator_name: str) -> dict:
    versions = list(
        db.scalars(
            select(SubmissionVersion)
            .where(SubmissionVersion.submission_id == submission.id)
            .order_by(SubmissionVersion.version.desc())
        )
    )
    return {
        "id": submission.id,
        "evaluatorName": evaluator_name,
        "status": submission.status,
        "score": float(submission.score),
        "responses": loads(submission.responses_json, {}),
        "raw": loads(submission.raw_payload_json, {}),
        "comments": submission.comments,
        "version": submission.version,
        "submittedAt": submission.submitted_at.isoformat() if submission.submitted_at else None,
        "history": [
            {
                "version": item.version,
                "payload": loads(item.payload_json, {}),
                "score": float(item.score),
                "actorType": item.actor_type,
                "actorName": item.actor_name,
                "reason": item.reason,
                "createdAt": item.created_at.isoformat(),
            }
            for item in versions
        ],
    }


def _admin_evaluation_payload(item: AdminEvaluation | None) -> dict | None:
    if not item:
        return None
    return {
        "id": item.id,
        "activityCode": item.activity_code,
        "score": float(item.score),
        "responses": loads(item.responses_json, {}),
        "raw": loads(item.raw_payload_json, {}),
        "comments": item.comments,
        "version": item.version,
        "updatedBy": item.updated_by,
        "updatedAt": item.updated_at.isoformat(),
    }


@router.get("/journeys/{journey_id}/recruits/{recruit_id}/admin-evaluations/{activity_code}")
def admin_evaluation_detail(
    journey_id: str,
    recruit_id: str,
    activity_code: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    journey = get_journey_or_404(db, journey_id)
    recruit = get_recruit_or_404(db, journey.id, recruit_id)
    if activity_code not in RUBRICS:
        raise HTTPException(status_code=404, detail="Unknown activity.")
    item = db.scalar(select(AdminEvaluation).where(
        AdminEvaluation.journey_id == journey.id,
        AdminEvaluation.recruit_id == recruit.id,
        AdminEvaluation.activity_code == activity_code,
    ))
    return {
        "recruit": serialize_recruit(recruit),
        "activityCode": activity_code,
        "activityName": RUBRICS[activity_code].name,
        "rubric": public_rubric(activity_code),
        "evaluation": _admin_evaluation_payload(item),
    }


@router.put("/journeys/{journey_id}/recruits/{recruit_id}/admin-evaluations/{activity_code}")
def save_admin_evaluation(
    journey_id: str,
    recruit_id: str,
    activity_code: str,
    payload: AdminEvaluationRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    recruit = get_recruit_or_404(db, journey.id, recruit_id)
    if activity_code not in RUBRICS:
        raise HTTPException(status_code=404, detail="Unknown activity.")
    item = db.scalar(select(AdminEvaluation).where(
        AdminEvaluation.journey_id == journey.id,
        AdminEvaluation.recruit_id == recruit.id,
        AdminEvaluation.activity_code == activity_code,
    ))
    if item and payload.client_version is not None and payload.client_version != item.version:
        raise HTTPException(status_code=409, detail="This admin evaluation changed elsewhere. Reload before saving.")
    score, normalized, normalized_raw = score_evaluation(activity_code, payload.responses, payload.raw)
    before = _admin_evaluation_payload(item)
    if not item:
        item = AdminEvaluation(
            journey_id=journey.id,
            recruit_id=recruit.id,
            activity_code=activity_code,
            updated_by=context.actor_name,
        )
        db.add(item)
    else:
        item.version += 1
    item.responses_json = dumps(normalized)
    item.raw_payload_json = dumps(normalized_raw)
    item.comments = payload.comments.strip()
    item.score = score
    item.updated_by = context.actor_name
    item.updated_at = utcnow()
    db.flush()
    after = _admin_evaluation_payload(item)
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="evaluation.admin_override_saved",
        entity_type="admin_evaluation",
        entity_id=item.id,
        before=before,
        after={**(after or {}), "recruitId": recruit.id, "recruitName": recruit.name},
        reason=payload.reason,
    )
    _commit(db)
    return {"ok": True, "evaluation": _admin_evaluation_payload(item)}


@router.delete("/journeys/{journey_id}/recruits/{recruit_id}/admin-evaluations/{activity_code}")
def delete_admin_evaluation(
    journey_id: str,
    recruit_id: str,
    activity_code: str,
    request: Request,
    reason: str = Query(min_length=1, max_length=1000),
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    recruit = get_recruit_or_404(db, journey.id, recruit_id)
    item = db.scalar(select(AdminEvaluation).where(
        AdminEvaluation.journey_id == journey.id,
        AdminEvaluation.recruit_id == recruit.id,
        AdminEvaluation.activity_code == activity_code,
    ))
    if not item:
        raise HTTPException(status_code=404, detail="Admin evaluation not found.")
    before = _admin_evaluation_payload(item)
    item_id = item.id
    db.delete(item)
    audit(db, journey_id=journey.id, actor_type="admin", actor_name=context.actor_name,
          action="evaluation.admin_override_removed", entity_type="admin_evaluation", entity_id=item_id,
          before=before, after={"recruitId": recruit.id, "recruitName": recruit.name, "activityCode": activity_code},
          reason=reason)
    _commit(db)
    return {"ok": True}


def _dimension_breakdowns(details: dict[str, list[dict]], result_row: dict | None) -> dict[str, dict]:
    """Build an admin-only, evaluator-attributed explanation of every dimension score."""
    result_dimensions = (result_row or {}).get("dimensions", {})
    result_activities = (result_row or {}).get("activities", {})
    breakdowns: dict[str, dict] = {}
    for dimension in DIMENSION_ORDER:
        if dimension in BEHAVIORAL_DIMENSIONS:
            activity_codes = DIMENSION_ACTIVITIES
        elif dimension == "application":
            activity_codes = ("skills",)
        else:
            activity_codes = ("sport",)

        matching_criteria = [
            (activity_code, criterion)
            for activity_code in activity_codes
            for criterion in RUBRICS[activity_code].criteria
            if dimension not in BEHAVIORAL_DIMENSIONS
            or criterion.dimension.casefold() == dimension
        ]
        total_weight = sum((criterion.weight for _activity_code, criterion in matching_criteria), Decimal("0"))
        activities: list[dict] = []
        for activity_code in activity_codes:
            criteria_payload: list[dict] = []
            for criterion_activity, criterion in matching_criteria:
                if criterion_activity != activity_code:
                    continue
                evaluator_payload: list[dict] = []
                grades: list[Decimal] = []
                for evaluation in details.get(activity_code, []):
                    submission = evaluation.get("submission")
                    final = submission and submission.get("status") in {"submitted", "locked"}
                    grade: Decimal | None = None
                    if final:
                        try:
                            candidate = Decimal(str(submission.get("responses", {}).get(criterion.key)))
                            if candidate.is_finite() and Decimal("0") <= candidate <= Decimal("5"):
                                grade = candidate
                                grades.append(candidate)
                        except (ArithmeticError, TypeError, ValueError):
                            grade = None
                    raw_value = submission.get("raw", {}).get(criterion.key) if submission else None
                    evaluator_payload.append(
                        {
                            "evaluatorId": evaluation["evaluatorId"],
                            "evaluatorName": evaluation["evaluatorName"],
                            "evaluatorRole": evaluation["evaluatorRole"],
                            "submissionId": submission.get("id") if submission else None,
                            "status": submission.get("status") if submission else "missing",
                            "grade": float(grade) if grade is not None else None,
                            "rawValue": raw_value,
                            "isAdmin": bool(evaluation.get("isAdmin")),
                        }
                    )
                admin_grades = [
                    Decimal(str(item["grade"]))
                    for item in evaluator_payload
                    if item.get("isAdmin") and item.get("grade") is not None
                ]
                if admin_grades:
                    grades = admin_grades
                average = sum(grades, Decimal("0")) / Decimal(len(grades)) if grades else None
                contribution = (
                    average / Decimal("5") * criterion.weight / total_weight
                    if average is not None and total_weight
                    else Decimal("0")
                )
                criteria_payload.append(
                    {
                        "key": criterion.key,
                        "name": criterion.name,
                        "explanation": criterion.explanation,
                        "weight": float(criterion.weight / total_weight) if total_weight else 0.0,
                        "criterionAverage": float(average) if average is not None else None,
                        "weightedContribution": float(contribution),
                        "unit": criterion.unit,
                        "evaluators": evaluator_payload,
                    }
                )
            activities.append(
                {
                    "code": activity_code,
                    "name": RUBRICS[activity_code].name,
                    "activityScore": result_activities.get(activity_code, {}).get("score", 0),
                    "criteria": criteria_payload,
                }
            )
        official = result_dimensions.get(dimension, {})
        breakdowns[dimension] = {
            "code": dimension,
            "name": DIMENSION_NAMES[dimension],
            "score": official.get("score", 0),
            "rank": official.get("rank"),
            "complete": official.get("complete", False),
            "activities": activities,
        }
    return breakdowns


@router.get("/journeys/{journey_id}/submissions/{submission_id}")
def submission_detail(
    journey_id: str,
    submission_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    get_journey_or_404(db, journey_id)
    submission = db.get(EvaluationSubmission, submission_id)
    if not submission or submission.journey_id != journey_id:
        raise HTTPException(status_code=404, detail="Evaluation submission not found.")
    evaluator = db.get(Evaluator, submission.evaluator_id)
    recruit = db.get(Recruit, submission.recruit_id)
    return {
        **_submission_payload(db, submission, evaluator.name if evaluator else "Unknown"),
        "activityCode": submission.activity_code,
        "activityName": RUBRICS[submission.activity_code].name,
        "recruitName": recruit.name if recruit else "Unknown",
        "rubric": public_rubric(submission.activity_code),
    }


@router.get("/journeys/{journey_id}/recruits/{recruit_id}/profile")
def recruit_profile(
    journey_id: str,
    recruit_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    journey = get_journey_or_404(db, journey_id)
    recruit = get_recruit_or_404(db, journey.id, recruit_id)
    results = result_snapshot(db, journey)
    result_row = next((item for item in results["rows"] if item["recruitId"] == recruit.id), None)
    if result_row is None:
        result_row = {
            "recruitId": recruit.id,
            "name": recruit.name,
            "activities": {code: {"score": 0.0, "expected": 0, "submitted": 0, "complete": False, "rank": None, "adminOverride": False} for code in ACTIVITY_ORDER},
            "dimensions": {code: {"name": DIMENSION_NAMES[code], "score": 0.0, "complete": False, "availableWeight": 0.0, "rank": None} for code in DIMENSION_ORDER},
            "generalAverage": 0.0,
            "generalComplete": False,
            "overallScore": 0.0,
            "overallRank": None,
            "color": "red",
            "missingCount": len(ACTIVITY_ORDER) + 3,
            "missingComponents": [RUBRICS[code].name for code in ACTIVITY_ORDER] + ["Punctuality", "Respect to us", "Seriousness"],
            "missingActivityCount": len(ACTIVITY_ORDER),
            "missingDimensionCount": len(DIMENSION_ORDER),
            "missingGeneralCount": 3,
            "complete": False,
        }
    assessment = db.get(GeneralAssessment, recruit.id)
    evaluators = {item.id: item for item in db.scalars(select(Evaluator).where(Evaluator.journey_id == journey.id))}
    states = list(db.scalars(select(ActivityState).where(ActivityState.journey_id == journey.id)))
    details: dict[str, list[dict]] = {code: [] for code in ACTIVITY_ORDER}
    for state in states:
        if not state.assignment_round_id:
            continue
        assignments = list(
            db.scalars(
                select(Assignment).where(
                    Assignment.round_id == state.assignment_round_id,
                    Assignment.recruit_id == recruit.id,
                )
            )
        )
        for assignment in assignments:
            submission = db.scalar(
                select(EvaluationSubmission).where(EvaluationSubmission.assignment_id == assignment.id)
            )
            evaluator = evaluators.get(assignment.evaluator_id)
            details[state.code].append(
                {
                    "assignmentId": assignment.id,
                    "slot": assignment.slot,
                    "roomNumber": assignment.room_number,
                    "evaluatorId": assignment.evaluator_id,
                    "evaluatorName": evaluator.name if evaluator else "Unknown",
                    "evaluatorRole": evaluator.role if evaluator else "",
                    "submission": _submission_payload(db, submission, evaluator.name if evaluator else "Unknown")
                    if submission
                    else None,
                }
            )
    admin_evaluations = list(db.scalars(select(AdminEvaluation).where(
        AdminEvaluation.journey_id == journey.id,
        AdminEvaluation.recruit_id == recruit.id,
    )))
    for item in admin_evaluations:
        details[item.activity_code].insert(0, {
            "assignmentId": None,
            "slot": 0,
            "roomNumber": None,
            "evaluatorId": None,
            "evaluatorName": f"Admin: {item.updated_by}",
            "evaluatorRole": "admin",
            "isAdmin": True,
            "submission": {
                **(_admin_evaluation_payload(item) or {}),
                "id": None,
                "status": "admin_override",
                "submittedAt": item.updated_at.isoformat(),
                "history": [],
            },
        })
    history = list(
        db.scalars(
            select(AuditEvent)
            .where(AuditEvent.journey_id == journey.id, AuditEvent.entity_id == recruit.id)
            .order_by(AuditEvent.created_at.desc())
        )
    )
    return {
        "recruit": serialize_recruit(recruit),
        "photoUrl": f"/api/admin/journeys/{journey.id}/recruits/{recruit.id}/photo" if recruit.photo_data else None,
        "result": result_row,
        "dimensionBreakdowns": _dimension_breakdowns(details, result_row),
        "dimensionAverages": results["dimensionAverages"],
        "activityAverages": results["activityAverages"],
        "assessment": {
            "punctuality": float(assessment.punctuality) if assessment and assessment.punctuality is not None else None,
            "respect": float(assessment.respect) if assessment and assessment.respect is not None else None,
            "seriousness": float(assessment.seriousness) if assessment and assessment.seriousness is not None else None,
            "comment": assessment.comment if assessment else "",
            "notes": assessment.notes if assessment else "",
            "version": assessment.version if assessment else 0,
        },
        "evaluations": details,
        "rubrics": {code: public_rubric(code) for code in ACTIVITY_ORDER},
        "adminEvaluations": {item.activity_code: _admin_evaluation_payload(item) for item in admin_evaluations},
        "history": [
            {
                "action": item.action,
                "actorName": item.actor_name,
                "reason": item.reason,
                "before": loads(item.before_json, {}),
                "after": loads(item.after_json, {}),
                "createdAt": item.created_at.isoformat(),
            }
            for item in history
        ],
    }


@router.put("/journeys/{journey_id}/recruits/{recruit_id}/profile")
def update_recruit_profile(
    journey_id: str,
    recruit_id: str,
    payload: GeneralAssessmentRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    recruit = get_recruit_or_404(db, journey.id, recruit_id)
    assessment = db.get(GeneralAssessment, recruit.id)
    if assessment and payload.base_version is not None and payload.base_version != assessment.version:
        raise HTTPException(status_code=409, detail="This profile was changed by another admin. Reload before saving.")
    before = {
        "punctuality": assessment.punctuality if assessment else None,
        "respect": assessment.respect if assessment else None,
        "seriousness": assessment.seriousness if assessment else None,
        "comment": assessment.comment if assessment else "",
        "notes": assessment.notes if assessment else "",
    }
    values: dict[str, Decimal | None] = {}
    for key, value in (
        ("punctuality", payload.punctuality),
        ("respect", payload.respect),
        ("seriousness", payload.seriousness),
    ):
        try:
            values[key] = None if value is None else validate_general_grade(value, key.title())
        except ScoringError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
    if not assessment:
        assessment = GeneralAssessment(recruit_id=recruit.id)
        db.add(assessment)
    else:
        assessment.version += 1
    assessment.punctuality = values["punctuality"]
    assessment.respect = values["respect"]
    assessment.seriousness = values["seriousness"]
    assessment.comment = payload.comment.strip()
    assessment.notes = payload.notes.strip()
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="recruit.profile_updated",
        entity_type="recruit",
        entity_id=recruit.id,
        before=before,
        after={**values, "comment": assessment.comment, "notes": assessment.notes},
    )
    _commit(db)
    return {"ok": True, "version": assessment.version}


@router.post("/journeys/{journey_id}/submissions/{submission_id}/correct")
def correct_submission(
    journey_id: str,
    submission_id: str,
    payload: AdminCorrectionRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    submission = db.get(EvaluationSubmission, submission_id)
    if not submission or submission.journey_id != journey.id:
        raise HTTPException(status_code=404, detail="Evaluation submission not found.")
    assignment = db.get(Assignment, submission.assignment_id)
    round_record = db.get(AssignmentRound, assignment.round_id) if assignment else None
    state = db.scalar(
        select(ActivityState).where(
            ActivityState.journey_id == journey.id,
            ActivityState.code == submission.activity_code,
        )
    )
    if not assignment or not round_record or not state:
        raise HTTPException(status_code=409, detail="The evaluation history is incomplete.")
    if payload.client_version is not None and payload.client_version != submission.version:
        raise HTTPException(status_code=409, detail="This evaluation was changed by another user. Reload before saving.")
    if not payload.reason.strip():
        raise HTTPException(status_code=422, detail="A correction reason is required.")
    saved = save_submission(
        db,
        assignment=assignment,
        round_record=round_record,
        state=state,
        responses=payload.responses,
        raw=payload.raw,
        comments=payload.comments,
        actor_type="admin",
        actor_name=context.actor_name,
        submit=True,
        reason=payload.reason,
    )
    if state.status == "closed":
        saved.status = "locked"
        version_record = db.scalar(
            select(SubmissionVersion).where(
                SubmissionVersion.submission_id == saved.id,
                SubmissionVersion.version == saved.version,
            )
        )
        if version_record:
            version_payload = loads(version_record.payload_json, {})
            version_payload["status"] = "locked"
            version_record.payload_json = dumps(version_payload)
    _commit(db)
    return {"ok": True, "version": saved.version, "score": float(saved.score), "status": saved.status}


@router.get("/journeys/{journey_id}/event-day-protection")
def event_day_protection_status(
    journey_id: str,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    get_journey_or_404(db, journey_id)
    protection = db.scalar(select(EventDayProtection).where(EventDayProtection.journey_id == journey_id))
    if protection:
        _sync_protection(db, protection)
        _commit(db)
    return _protection_payload(protection)


@router.post("/journeys/{journey_id}/event-day-protection/start")
def start_event_day_protection(
    journey_id: str,
    payload: EventDayProtectionStartRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    if not is_configured():
        raise HTTPException(status_code=503, detail="Event-day protection is not configured on this server.")
    journey = get_journey_or_404(db, journey_id)
    if journey.status in {"completed", "archived"}:
        raise HTTPException(status_code=409, detail="Reopen this Journee before starting event-day protection.")
    now = utcnow()
    other = db.scalar(
        select(EventDayProtection).where(
            EventDayProtection.journey_id != journey.id,
            EventDayProtection.status.in_(["starting", "active", "incident"]),
            EventDayProtection.ends_at > now,
        )
    )
    if other:
        other_journey = db.get(Journey, other.journey_id)
        raise HTTPException(
            status_code=409,
            detail=f"Protection is already active for {other_journey.name if other_journey else 'another Journee'}.",
        )
    existing_active = db.scalar(select(Journey).where(Journey.status == "active", Journey.id != journey.id))
    if existing_active:
        raise HTTPException(status_code=409, detail=f"{existing_active.name} is already the active Journee.")
    protection = db.scalar(select(EventDayProtection).where(EventDayProtection.journey_id == journey.id))
    if protection and protection.status in {"starting", "active"} and _aware(protection.ends_at) > now:
        raise HTTPException(status_code=409, detail="Event-day protection is already running for this Journee.")
    activation_id = new_id()
    if not protection:
        protection = EventDayProtection(
            journey_id=journey.id,
            activation_id=activation_id,
            duration_hours=payload.duration_hours,
            status="starting",
            started_at=now,
            ends_at=now + timedelta(hours=payload.duration_hours),
        )
        db.add(protection)
    else:
        protection.activation_id = activation_id
        protection.duration_hours = payload.duration_hours
        protection.status = "starting"
        protection.github_run_id = None
        protection.github_run_url = None
        protection.last_error = ""
        protection.started_at = now
        protection.ends_at = now + timedelta(hours=payload.duration_hours)
        protection.stopped_at = None
        protection.version += 1
    db.flush()
    before_status = journey.status
    journey.status = "active"
    journey.archived_at = None
    journey.version += 1
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="event_day_protection.started",
        entity_type="event_day_protection",
        entity_id=protection.id,
        before={"journeyStatus": before_status},
        after={"durationHours": payload.duration_hours, "activationId": activation_id},
    )
    _commit(db)
    try:
        dispatch_monitor(payload.duration_hours, activation_id)
    except ProtectionControlError as exc:
        protection = db.scalar(select(EventDayProtection).where(EventDayProtection.journey_id == journey.id))
        protection.status = "incident"
        protection.last_error = str(exc)
        protection.version += 1
        _commit(db)
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    return _protection_payload(protection)


@router.post("/journeys/{journey_id}/event-day-protection/restart")
def restart_event_day_protection(
    journey_id: str,
    payload: EventDayProtectionActionRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    protection = db.scalar(select(EventDayProtection).where(EventDayProtection.journey_id == journey.id))
    if not protection:
        raise HTTPException(status_code=409, detail="Start event-day protection first.")
    if protection.github_run_id or protection.activation_id:
        try:
            _cancel_protection_monitor(protection)
        except ProtectionControlError:
            pass
    now = utcnow()
    activation_id = new_id()
    protection.activation_id = activation_id
    protection.status = "starting"
    protection.github_run_id = None
    protection.github_run_url = None
    protection.last_error = ""
    protection.started_at = now
    protection.ends_at = now + timedelta(hours=protection.duration_hours)
    protection.stopped_at = None
    protection.version += 1
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="event_day_protection.restarted",
        entity_type="event_day_protection",
        entity_id=protection.id,
        after={"durationHours": protection.duration_hours, "activationId": activation_id},
        reason=payload.reason or "Admin restarted event-day protection.",
    )
    _commit(db)
    try:
        dispatch_monitor(protection.duration_hours, activation_id)
    except ProtectionControlError as exc:
        protection.status = "incident"
        protection.last_error = str(exc)
        protection.version += 1
        _commit(db)
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    return _protection_payload(protection)


@router.post("/journeys/{journey_id}/event-day-protection/end")
def end_event_day(
    journey_id: str,
    payload: EventDayProtectionActionRequest,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    protection = db.scalar(select(EventDayProtection).where(EventDayProtection.journey_id == journey.id))
    if protection:
        try:
            _cancel_protection_monitor(protection)
        except ProtectionControlError as exc:
            protection.last_error = str(exc)
    now = utcnow()
    if protection:
        protection.status = "stopped"
        protection.stopped_at = now
        protection.version += 1
    for activity in db.scalars(
        select(ActivityState).where(ActivityState.journey_id == journey.id, ActivityState.status == "open")
    ):
        activity.status = "closed"
        activity.closed_at = now
        activity.version += 1
    before_status = journey.status
    journey.status = "completed"
    journey.current_activity = None
    journey.version += 1
    audit(
        db,
        journey_id=journey.id,
        actor_type="admin",
        actor_name=context.actor_name,
        action="journey.completed_with_protection_stopped",
        entity_type="journey",
        entity_id=journey.id,
        before={"status": before_status},
        after={"status": "completed", "protectionStatus": "stopped"},
        reason=payload.reason or "Journee completed from event-day control.",
    )
    _commit(db)
    return {"ok": True, "protection": _protection_payload(protection), "journeyStatus": journey.status}


@router.get("/journeys/{journey_id}/audit")
def audit_log(
    journey_id: str,
    limit: int = Query(default=500, ge=1, le=5000),
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    del context
    get_journey_or_404(db, journey_id)
    events = list(
        db.scalars(
            select(AuditEvent)
            .where(AuditEvent.journey_id == journey_id)
            .order_by(AuditEvent.created_at.desc())
            .limit(limit)
        )
    )
    return [
        {
            "id": item.id,
            "actorType": item.actor_type,
            "actorName": item.actor_name,
            "action": item.action,
            "entityType": item.entity_type,
            "entityId": item.entity_id,
            "before": loads(item.before_json, {}),
            "after": loads(item.after_json, {}),
            "reason": item.reason,
            "createdAt": item.created_at.isoformat(),
        }
        for item in events
    ]


@router.delete("/journeys/{journey_id}")
def delete_empty_draft(
    journey_id: str,
    request: Request,
    context: AdminContext = Depends(require_admin),
    db: Session = Depends(get_db),
):
    require_csrf(request, context.csrf_token)
    journey = get_journey_or_404(db, journey_id)
    audit(db, journey_id=None, actor_type="admin", actor_name=context.actor_name,
          action="journey.deleted", entity_type="journey", entity_id=journey.id,
          before=serialize_journey(db, journey), reason="Permanent deletion requested by admin.")
    db.delete(journey)
    _commit(db)
    return {"ok": True}
