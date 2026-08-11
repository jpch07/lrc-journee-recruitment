from __future__ import annotations

import io
import zipfile
from datetime import date

from openpyxl import load_workbook
from sqlalchemy import func, select

from app.db import SessionLocal
from app.models import (ActivityState, AdminEvaluation, Assignment, AssignmentRound, EvaluationSubmission,
                        Evaluator, GeneralAssessment, Journey, Recruit, SubmissionVersion)
from app.rubric import ACTIVITY_ORDER, DIMENSION_ORDER, RUBRICS
from app.services import create_journey, result_snapshot
from app.utils import dumps


def admin_login(client):
    response = client.post("/api/admin/login", json={"password": "test-password", "display_name": "Test Admin"})
    assert response.status_code == 200, response.text
    return response.json()["csrfToken"]


def admin_headers(csrf):
    return {"X-CSRF-Token": csrf}


def test_management_view_defaults_to_completed_journees_and_home_has_three_destinations(client):
    login = client.post("/api/auth/login", json={"username": "JP Chaaya", "password": "test-password"})
    assert login.status_code == 200, login.text
    with SessionLocal() as db:
        first = create_journey(db, "Completed One", date(2026, 8, 1), 1, "Test")
        second = create_journey(db, "Completed Two", date(2026, 8, 2), 1, "Test")
        draft = create_journey(db, "Draft Three", date(2026, 8, 3), 1, "Test")
        first.status = "completed"
        second.status = "completed"
        db.add_all([
            Recruit(journey_id=first.id, name="Alpha", present=True),
            Recruit(journey_id=second.id, name="Bravo", present=True),
            Recruit(journey_id=draft.id, name="Charlie", present=True),
        ])
        db.commit()

    aggregate = client.get("/api/view/completed")
    assert aggregate.status_code == 200, aggregate.text
    payload = aggregate.json()
    assert payload["scope"] == "completed"
    assert {item["journeyName"] for item in payload["recruits"]} == {"Completed One", "Completed Two"}
    assert {item["name"] for item in payload["results"]["rows"]} == {"Alpha", "Bravo"}
    assert all(item["overallRank"] == 1 for item in payload["results"]["rows"])
    assert all(item["profileKey"] for item in payload["results"]["rows"])

    report = client.get("/api/view/report.xlsx")
    assert report.status_code == 200, report.text
    with zipfile.ZipFile(io.BytesIO(report.content)) as archive:
        names = set(archive.namelist())
        assert "xl/metadata.xml" in names
        assert "xl/richData/rdrichvalue.xml" in names
        assert any(name.startswith("xl/media/profile-photo-") for name in names)
        profile_xml = next(
            archive.read(name).decode("utf-8")
            for name in names
            if name.startswith("xl/worksheets/sheet") and b"Recruit profile" in archive.read(name)
        )
        assert 'r="J3"' in profile_xml and 'vm="' in profile_xml
    workbook = load_workbook(io.BytesIO(report.content), read_only=False, data_only=False)
    assert workbook.sheetnames == ["Attendance", "Results", "Recruit Profiles"]
    report_values = {
        value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if isinstance(value, str)
    }
    assert "Draft Three" not in report_values
    assert "Charlie" not in report_values
    attendance_headers = [workbook["Attendance"].cell(5, column).value for column in range(1, 8)]
    assert attendance_headers == ["Journee", "Recruit", "Phone number", "Date of birth", "Status", "Arrival time", "Attendance comment"]
    assert "Roster" not in report_values
    profile_options = [workbook["Recruit Profiles"].cell(row, 90).value for row in range(2, 4)]
    assert profile_options == ["Alpha", "Bravo"]
    assert "XLOOKUP" in workbook["Recruit Profiles"]["J3"].value
    assert "J3:L7" in {str(item) for item in workbook["Recruit Profiles"].merged_cells.ranges}
    workbook.close()

    home = client.get("/")
    assert home.status_code == 200
    assert all(path in home.text for path in ('href="/admin"', 'href="/evaluate"', 'href="/view"'))


def test_completion_lists_missing_activities_and_general_grades_not_dimensions():
    with SessionLocal() as db:
        journey = create_journey(db, "Incomplete Journey", date(2026, 9, 1), 1, "Test")
        recruit = Recruit(journey_id=journey.id, name="Incomplete Recruit", present=True)
        db.add(recruit)
        db.flush()
        row = result_snapshot(db, journey)["rows"][0]

    assert row["missingCount"] == 8
    assert row["missingActivityCount"] == 5
    assert row["missingGeneralCount"] == 3
    assert row["missingComponents"] == [
        *[RUBRICS[code].name for code in ACTIVITY_ORDER],
        "Punctuality", "Respect to us", "Seriousness",
    ]
    assert not any(name in row["missingComponents"] for name in ("Willingness", "Adaptability", "Intelligence"))


def test_profile_notes_are_saved_audited_and_exported(client):
    csrf = admin_login(client)
    journey = client.post(
        "/api/admin/journeys",
        headers=admin_headers(csrf),
        json={"name": "Notes Journee", "event_date": "2026-08-11", "room_count": 1},
    ).json()
    recruit = client.post(
        f"/api/admin/journeys/{journey['id']}/recruits",
        headers=admin_headers(csrf),
        json={"name": "Recruit Notes"},
    ).json()

    profile_url = f"/api/admin/journeys/{journey['id']}/recruits/{recruit['id']}/profile"
    profile = client.get(profile_url).json()
    assert profile["assessment"]["notes"] == ""

    response = client.put(
        profile_url,
        headers=admin_headers(csrf),
        json={
            "punctuality": 0.8,
            "respect": 0.9,
            "seriousness": 1.0,
            "comment": "General comment",
            "notes": "Private follow-up note",
            "base_version": 0,
        },
    )
    assert response.status_code == 200, response.text

    profile = client.get(profile_url).json()
    assert profile["assessment"]["notes"] == "Private follow-up note"
    assert profile["history"][0]["after"]["notes"] == "Private follow-up note"
    with SessionLocal() as db:
        saved_recruit = db.get(Recruit, recruit["id"])
        saved_recruit.present = True
        db.commit()
    results = client.get(f"/api/admin/journeys/{journey['id']}/results").json()
    result_row = next(item for item in results["rows"] if item["recruitId"] == recruit["id"])
    assert result_row["generalComment"] == "General comment"
    assert result_row["notes"] == "Private follow-up note"

    with SessionLocal() as db:
        saved_journey = db.get(Journey, journey["id"])
        saved_journey.status = "completed"
        db.commit()

    exported = client.get(f"/api/admin/journeys/{journey['id']}/export.xlsx")
    assert exported.headers["cache-control"] == "no-store"
    workbook = load_workbook(io.BytesIO(exported.content), read_only=False, data_only=False)
    assert workbook.sheetnames == ["Attendance", "Results", "Recruit Profiles"]
    profile_sheet = workbook["Recruit Profiles"]
    values = [cell for row in profile_sheet.iter_rows(values_only=True) for cell in row if cell is not None]
    assert "Private follow-up note" in values
    results_sheet = workbook["Results"]
    assert [results_sheet.cell(5, column).value for column in (9, 10)] == ["General comment", "Notes"]
    result_values = [cell for row in results_sheet.iter_rows(values_only=True) for cell in row if cell is not None]
    assert "General comment" in result_values
    assert "Private follow-up note" in result_values
    assert "_xlfn.XLOOKUP" in results_sheet["I6"].value
    workbook.close()


def test_shared_skills_simulation_assignment_is_not_a_warning(client):
    csrf = admin_login(client)
    with SessionLocal() as db:
        journey = create_journey(db, "Shared Pairing", date(2026, 8, 12), 1, "Test Admin")
        db.flush()
        simulation_round = AssignmentRound(
            journey_id=journey.id,
            activity_code="simulation",
            version=1,
            status="published",
            seed="shared-seed",
            warnings_json=dumps(["Shared Skills & Simulation assignment."]),
            created_by="Test Admin",
        )
        db.add(simulation_round)
        db.flush()
        state = db.scalar(
            select(ActivityState).where(
                ActivityState.journey_id == journey.id,
                ActivityState.code == "simulation",
            )
        )
        state.assignment_round_id = simulation_round.id
        journey_id = journey.id
        db.commit()

    dashboard = client.get(f"/api/admin/journeys/{journey_id}/dashboard")
    assert dashboard.status_code == 200, dashboard.text
    assert dashboard.json()["warnings"] == []
    assignment = client.get(f"/api/admin/journeys/{journey_id}/assignments/simulation")
    assert assignment.status_code == 200, assignment.text
    assert assignment.json()["warnings"] == []


def test_complete_sport_workflow_and_isolation(client):
    csrf = admin_login(client)
    response = client.post(
        "/api/admin/journeys",
        headers=admin_headers(csrf),
        json={"name": "Pilot Journee", "event_date": "2026-08-10", "room_count": 1},
    )
    assert response.status_code == 200, response.text
    journey = response.json()

    recruits = []
    for index in range(2):
        response = client.post(
            f"/api/admin/journeys/{journey['id']}/recruits",
            headers=admin_headers(csrf),
            json={"name": f"Recruit {index + 1}", "phone_number": f"701000{index + 1}",
                  "date_of_birth": f"200{index}-01-02"},
        )
        recruits.append(response.json())
    assert recruits[0]["dateOfBirth"] == "2000-01-02"
    evaluators = []
    for index, role in enumerate(("overall", "dossard")):
        response = client.post(
            f"/api/admin/journeys/{journey['id']}/evaluators",
            headers=admin_headers(csrf),
                json={"name": f"Evaluator {index + 1}", "role": role, "add_to_directory": True, "password": f"temporary-pass-{index}"},
        )
        evaluators.append(response.json())

    response = client.put(
        f"/api/admin/journeys/{journey['id']}/attendance/recruits",
        headers=admin_headers(csrf),
        json={"items": [{"id": item["id"], "present": True, "arrival_time": "2026-08-10T07:00:00+03:00", "active": True, "base_version": item["version"]} for item in recruits]},
    )
    assert response.status_code == 200, response.text
    detail = client.get(f"/api/admin/journeys/{journey['id']}").json()
    assert next(item for item in detail["recruits"] if item["id"] == recruits[0]["id"])["dateOfBirth"] == "2000-01-02"
    with SessionLocal() as db:
        saved_journey = db.get(Journey, journey["id"])
        saved_journey.status = "completed"
        db.commit()
    exported = client.get(f"/api/admin/journeys/{journey['id']}/export.xlsx")
    assert exported.status_code == 200
    workbook = load_workbook(io.BytesIO(exported.content), read_only=True, data_only=True)
    assert workbook.sheetnames == ["Attendance", "Results", "Recruit Profiles"]
    recruit_rows = list(workbook["Attendance"].iter_rows(min_row=5, max_row=5, values_only=True))
    assert "Date of birth" in recruit_rows[0]
    attendance_values = [cell for row in workbook["Attendance"].iter_rows(values_only=True) for cell in row if cell is not None]
    assert any(getattr(value, "isoformat", lambda: "")().startswith("2000-01-02") for value in attendance_values)
    workbook.close()
    with SessionLocal() as db:
        saved_journey = db.get(Journey, journey["id"])
        saved_journey.status = "draft"
        db.commit()
    response = client.put(
        f"/api/admin/journeys/{journey['id']}/attendance/evaluators",
        headers=admin_headers(csrf),
        json={"items": [{"id": item["id"], "present": True, "role": item["role"], "active": True, "base_version": item["version"]} for item in evaluators]},
    )
    assert response.status_code == 200, response.text

    preview = client.post(
        f"/api/admin/journeys/{journey['id']}/assignments/sport/preview",
        headers=admin_headers(csrf),
        json={"seed": "sport-seed"},
    ).json()
    assert len(preview["assignments"]) == 2
    response = client.post(
        f"/api/admin/journeys/{journey['id']}/assignments/{preview['id']}/publish",
        headers=admin_headers(csrf),
    )
    assert response.status_code == 200, response.text
    response = client.post(
        f"/api/admin/journeys/{journey['id']}/activities/sport/open",
        headers=admin_headers(csrf),
        json={"reason": ""},
    )
    assert response.status_code == 200, response.text

    evaluator = evaluators[0]
    response = client.get(f"/api/public/journeys/{journey['publicToken']}")
    assert {item["id"] for item in response.json()["evaluators"]} == {item["id"] for item in evaluators}
    current = client.get("/api/public/current")
    assert current.status_code == 200
    assert current.json()["name"] == "Pilot Journee"
    response = client.post(
        "/api/public/current/session",
        json={"evaluator_id": evaluator["id"]},
    )
    evaluator_csrf = response.json()["csrfToken"]
    home = client.get("/api/evaluator/home").json()
    sport = next(item for item in home["activities"] if item["code"] == "sport")
    assert all("weight" not in criterion and "target" not in criterion
               for activity in home["activities"] for criterion in activity["rubric"]["criteria"])
    assert len(sport["tasks"]) == 1
    task = sport["tasks"][0]
    other_assignment = next(item for item in preview["assignments"] if item["id"] != task["assignmentId"])
    assert client.get(f"/api/evaluator/tasks/{other_assignment['id']}/photo").status_code == 404

    response = client.put(
        f"/api/evaluator/tasks/{task['assignmentId']}/draft",
        headers={"X-CSRF-Token": evaluator_csrf, "Idempotency-Key": "draft-one"},
        json={"responses": {}, "raw": {"push_ups": "10"}, "comments": "starting"},
    )
    assert response.status_code == 200, response.text
    draft_version = response.json()["submission"]["version"]
    payload = {
        "responses": {},
        "raw": {"push_ups": "30", "wall_sit": "120s", "beep_test": "10:00", "plank": "2m"},
        "comments": "complete",
        "client_version": draft_version,
    }
    headers = {"X-CSRF-Token": evaluator_csrf, "Idempotency-Key": "submit-once"}
    response = client.post(f"/api/evaluator/tasks/{task['assignmentId']}/submit", headers=headers, json=payload)
    assert response.status_code == 200, response.text
    assert response.json()["submission"]["score"] == 5.0
    submitted = response.json()["submission"]
    duplicate = client.post(f"/api/evaluator/tasks/{task['assignmentId']}/submit", headers=headers, json=payload)
    assert duplicate.status_code == 200
    with SessionLocal() as db:
        submission_count = db.scalar(select(func.count()).select_from(EvaluationSubmission))
        version_count = db.scalar(select(func.count()).select_from(SubmissionVersion))
    assert submission_count == 1
    assert version_count == 2

    results = client.get(f"/api/admin/journeys/{journey['id']}/results").json()
    evaluated = next(row for row in results["rows"] if row["recruitId"] == task["recruitId"])
    assert evaluated["activities"]["sport"]["score"] == 5.0
    assert evaluated["activities"]["sport"]["complete"] is True
    profile = client.get(
        f"/api/admin/journeys/{journey['id']}/recruits/{task['recruitId']}/profile"
    ).json()
    physical = profile["dimensionBreakdowns"]["physical_ability"]
    assert physical["score"] == 1.0
    assert physical["activities"][0]["code"] == "sport"
    assert len(physical["activities"][0]["criteria"]) == 4
    push_ups = next(item for item in physical["activities"][0]["criteria"] if item["key"] == "push_ups")
    assert push_ups["criterionAverage"] == 5.0
    assert push_ups["evaluators"][0]["evaluatorName"] == evaluator["name"]
    assert push_ups["evaluators"][0]["grade"] == 5.0
    assert push_ups["evaluators"][0]["rawValue"] == 30.0

    correction_url = f"/api/admin/journeys/{journey['id']}/submissions/{submitted['id']}/correct"
    corrected_payload = {
        "responses": {},
        "raw": {"push_ups": 15, "wall_sit": 60, "beep_test": 300, "plank": 60},
        "comments": "Admin verified result",
        "reason": "Corrected from the signed sport sheet",
        "client_version": submitted["version"],
    }
    corrected = client.post(correction_url, headers=admin_headers(csrf), json=corrected_payload)
    assert corrected.status_code == 200, corrected.text
    assert corrected.json()["score"] == 2.5
    assert corrected.json()["status"] == "submitted"
    stale = client.post(correction_url, headers=admin_headers(csrf), json=corrected_payload)
    assert stale.status_code == 409

    results_export = client.get(f"/api/admin/journeys/{journey['id']}/results.xlsx")
    assert results_export.status_code == 200
    result_workbook = load_workbook(io.BytesIO(results_export.content), read_only=False, data_only=False)
    assert result_workbook.sheetnames == ["Attendance", "Results", "Recruit Profiles"]
    assert "_xlfn.XLOOKUP" in result_workbook["Attendance"]["A6"].value
    assert "_xlfn.XLOOKUP" in result_workbook["Results"]["A6"].value
    profile_sheet = result_workbook["Recruit Profiles"]
    assert "_xlfn.XLOOKUP" in profile_sheet["H3"].value
    assert len(profile_sheet._charts) == 2
    profile_values = [cell for row in profile_sheet.iter_rows(values_only=True) for cell in row if cell is not None]
    assert "Criterion-level grading" in profile_values
    result_workbook.close()

    response = client.post(
        f"/api/admin/journeys/{journey['id']}/activities/sport/close",
        headers=admin_headers(csrf),
        json={"reason": ""},
    )
    assert response.status_code == 200
    closed_correction = client.post(
        correction_url,
        headers=admin_headers(csrf),
        json={**payload, "reason": "Final admin verification after closure", "client_version": corrected.json()["version"]},
    )
    assert closed_correction.status_code == 200, closed_correction.text
    assert closed_correction.json()["score"] == 5.0
    assert closed_correction.json()["status"] == "locked"
    blocked = client.put(
        f"/api/evaluator/tasks/{task['assignmentId']}/draft",
        headers={"X-CSRF-Token": evaluator_csrf},
        json={"responses": {}, "raw": {"push_ups": "31"}, "comments": "late"},
    )
    assert blocked.status_code == 409


def test_only_one_journey_can_be_active(client):
    csrf = admin_login(client)
    first = client.post("/api/admin/journeys", headers=admin_headers(csrf),
                        json={"name": "First", "event_date": "2026-09-01", "room_count": 1}).json()
    second = client.post("/api/admin/journeys", headers=admin_headers(csrf),
                         json={"name": "Second", "event_date": "2026-09-02", "room_count": 1}).json()
    assert client.patch(f"/api/admin/journeys/{first['id']}", headers=admin_headers(csrf),
                        json={"status": "active", "base_version": first["version"]}).status_code == 200
    blocked = client.patch(f"/api/admin/journeys/{second['id']}", headers=admin_headers(csrf),
                           json={"status": "active", "base_version": second["version"]})
    assert blocked.status_code == 409
    assert client.get("/api/public/current").json()["name"] == "First"


def test_recruit_import_accepts_phone_and_date_of_birth(client):
    csrf = admin_login(client)
    journey = client.post(
        "/api/admin/journeys",
        headers=admin_headers(csrf),
        json={"name": "DOB Import", "event_date": "2026-09-07"},
    ).json()
    response = client.post(
        f"/api/admin/journeys/{journey['id']}/import/recruits",
        headers=admin_headers(csrf),
        files={"file": ("recruits.csv", b"Name,Phone Number,Date of Birth\nImported Recruit,70123456,03/04/2001\n", "text/csv")},
    )
    assert response.status_code == 200, response.text
    assert response.json()["created"][0]["phoneNumber"] == "70123456"
    assert response.json()["created"][0]["dateOfBirth"] == "2001-04-03"


def test_received_submission_is_not_diluted_by_missing_co_evaluator():
    with SessionLocal() as db:
        journey = create_journey(db, "Partial Average", date(2026, 9, 3), 1, "Test")
        recruit = Recruit(journey_id=journey.id, name="Recruit", present=True)
        first = Evaluator(journey_id=journey.id, name="First", role="overall", present=True)
        second = Evaluator(journey_id=journey.id, name="Second", role="dossard", present=True)
        db.add_all([recruit, first, second]); db.flush()
        round_record = AssignmentRound(journey_id=journey.id, activity_code="sport", version=1,
                                       status="published", seed="partial", created_by="Test")
        db.add(round_record); db.flush()
        one = Assignment(round_id=round_record.id, evaluator_id=first.id, recruit_id=recruit.id, slot=1)
        two = Assignment(round_id=round_record.id, evaluator_id=second.id, recruit_id=recruit.id, slot=2)
        db.add_all([one, two]); db.flush()
        db.add(EvaluationSubmission(assignment_id=one.id, journey_id=journey.id, activity_code="sport",
                                    evaluator_id=first.id, recruit_id=recruit.id, status="locked", score=4))
        state = db.scalar(select(ActivityState).where(ActivityState.journey_id == journey.id,
                                                       ActivityState.code == "sport"))
        state.assignment_round_id = round_record.id
        db.flush()
        row = result_snapshot(db, journey)["rows"][0]
        assert row["activities"]["sport"]["score"] == 4.0
        assert row["activities"]["sport"]["submitted"] == 1
        assert row["activities"]["sport"]["expected"] == 2


def test_admin_evaluation_replaces_evaluator_average_and_dimension_inputs():
    with SessionLocal() as db:
        journey = create_journey(db, "Admin Override", date(2026, 9, 8), 1, "Test")
        recruit = Recruit(journey_id=journey.id, name="Recruit", present=True)
        evaluator = Evaluator(journey_id=journey.id, name="Evaluator", role="overall", present=True)
        db.add_all([recruit, evaluator]); db.flush()
        round_record = AssignmentRound(journey_id=journey.id, activity_code="escape_room", version=1,
                                       status="published", seed="override", created_by="Test")
        db.add(round_record); db.flush()
        assignment = Assignment(round_id=round_record.id, evaluator_id=evaluator.id, recruit_id=recruit.id, slot=1)
        db.add(assignment); db.flush()
        evaluator_responses = {criterion.key: 1 for criterion in RUBRICS["escape_room"].criteria}
        admin_responses = {criterion.key: 5 for criterion in RUBRICS["escape_room"].criteria}
        db.add(EvaluationSubmission(assignment_id=assignment.id, journey_id=journey.id,
                                    activity_code="escape_room", evaluator_id=evaluator.id,
                                    recruit_id=recruit.id, status="locked", score=1,
                                    responses_json=dumps(evaluator_responses)))
        db.add(AdminEvaluation(journey_id=journey.id, recruit_id=recruit.id,
                               activity_code="escape_room", score=5,
                               responses_json=dumps(admin_responses), updated_by="JP Chaaya"))
        state = db.scalar(select(ActivityState).where(ActivityState.journey_id == journey.id,
                                                       ActivityState.code == "escape_room"))
        state.assignment_round_id = round_record.id
        db.flush()
        row = result_snapshot(db, journey)["rows"][0]
        assert row["activities"]["escape_room"]["score"] == 5.0
        assert row["activities"]["escape_room"]["adminOverride"] is True
        assert row["dimensions"]["willingness"]["score"] == 0.2


def test_admin_can_create_an_evaluation_without_an_assignment(client):
    csrf = admin_login(client)
    headers = admin_headers(csrf)
    journey = client.post("/api/admin/journeys", headers=headers,
                          json={"name": "Manual Evaluation", "event_date": "2026-09-10"}).json()
    recruit = client.post(f"/api/admin/journeys/{journey['id']}/recruits", headers=headers,
                          json={"name": "Unassigned Recruit"}).json()
    responses = {criterion.key: 4 for criterion in RUBRICS["negotiation"].criteria}
    saved = client.put(
        f"/api/admin/journeys/{journey['id']}/recruits/{recruit['id']}/admin-evaluations/negotiation",
        headers=headers,
        json={"responses": responses, "raw": {}, "comments": "Entered by administration",
              "reason": "Paper form entered manually"},
    )
    assert saved.status_code == 200, saved.text
    detail = client.get(
        f"/api/admin/journeys/{journey['id']}/recruits/{recruit['id']}/admin-evaluations/negotiation"
    ).json()
    assert detail["evaluation"]["score"] == 4.0


def test_six_dimension_results_reach_exact_overall_maximum():
    with SessionLocal() as db:
        journey = create_journey(db, "Dimension Maximum", date(2026, 9, 6), 1, "Test")
        recruit = Recruit(journey_id=journey.id, name="Complete Recruit", present=True)
        evaluator = Evaluator(journey_id=journey.id, name="Complete Evaluator", role="overall", present=True)
        db.add_all([recruit, evaluator]); db.flush()
        states = {item.code: item for item in db.scalars(select(ActivityState).where(
            ActivityState.journey_id == journey.id))}
        for code in ACTIVITY_ORDER:
            round_record = AssignmentRound(journey_id=journey.id, activity_code=code, version=1,
                                           status="published", seed=code, created_by="Test")
            db.add(round_record); db.flush()
            assignment = Assignment(round_id=round_record.id, evaluator_id=evaluator.id,
                                    recruit_id=recruit.id, slot=1)
            db.add(assignment); db.flush()
            responses = ({criterion.key: 5 for criterion in RUBRICS[code].criteria}
                         if RUBRICS[code].kind == "weighted" else {})
            db.add(EvaluationSubmission(
                assignment_id=assignment.id,
                journey_id=journey.id,
                activity_code=code,
                evaluator_id=evaluator.id,
                recruit_id=recruit.id,
                status="locked",
                score=5,
                responses_json=dumps(responses),
            ))
            states[code].assignment_round_id = round_record.id
        db.add(GeneralAssessment(recruit_id=recruit.id, punctuality=1, respect=1, seriousness=1))
        db.flush()

        snapshot = result_snapshot(db, journey)
        row = snapshot["rows"][0]
        assert set(row["dimensions"]) == set(DIMENSION_ORDER)
        assert all(item["score"] == 1.0 for item in row["dimensions"].values())
        assert all(item["complete"] is True for item in row["dimensions"].values())
        assert row["overallScore"] == 20.0
        assert row["missingCount"] == 0
        assert all(value == 1.0 for value in snapshot["dimensionAverages"].values())


def test_skills_and_simulation_can_open_and_close_together(client):
    csrf = admin_login(client)
    journey = client.post("/api/admin/journeys", headers=admin_headers(csrf),
                          json={"name": "Paired Activities", "event_date": "2026-09-04"}).json()
    with SessionLocal() as db:
        states = {item.code: item for item in db.scalars(select(ActivityState).where(
            ActivityState.journey_id == journey["id"]))}
        states["negotiation"].status = "closed"
        for code in ("skills", "simulation"):
            round_record = AssignmentRound(journey_id=journey["id"], activity_code=code, version=1,
                                           status="published", seed=code, created_by="Test")
            db.add(round_record); db.flush(); states[code].assignment_round_id = round_record.id
        db.commit()
    opened = client.post(f"/api/admin/journeys/{journey['id']}/activities/skills-simulation/open",
                         headers=admin_headers(csrf), json={"reason": ""})
    assert opened.status_code == 200, opened.text
    with SessionLocal() as db:
        statuses = {item.code: item.status for item in db.scalars(select(ActivityState).where(
            ActivityState.journey_id == journey["id"], ActivityState.code.in_(("skills", "simulation"))))}
        assert statuses == {"skills": "open", "simulation": "open"}
    closed = client.post(f"/api/admin/journeys/{journey['id']}/activities/skills-simulation/close",
                         headers=admin_headers(csrf), json={"reason": ""})
    assert closed.status_code == 200, closed.text


def test_room_count_moves_to_assignments_and_journey_with_data_can_be_deleted(client):
    csrf = admin_login(client)
    journey = client.post("/api/admin/journeys", headers=admin_headers(csrf),
                          json={"name": "Delete Me", "event_date": "2026-09-05"}).json()
    changed = client.put(f"/api/admin/journeys/{journey['id']}/rooms/count", headers=admin_headers(csrf),
                         json={"room_count": 3, "reason": ""})
    assert changed.status_code == 200
    recruit = client.post(f"/api/admin/journeys/{journey['id']}/recruits", headers=admin_headers(csrf),
                          json={"name": "Temporary"}).json()
    assert client.delete(f"/api/admin/journeys/{journey['id']}/recruits/{recruit['id']}",
                         headers=admin_headers(csrf)).json()["disposition"] == "deleted"
    assert client.delete(f"/api/admin/journeys/{journey['id']}", headers=admin_headers(csrf)).status_code == 200
    with SessionLocal() as db:
        assert db.get(Journey, journey["id"]) is None


def test_admin_stale_attendance_is_rejected(client):
    csrf = admin_login(client)
    journey = client.post(
        "/api/admin/journeys",
        headers=admin_headers(csrf),
        json={"name": "Conflict", "event_date": "2026-09-01", "room_count": 1},
    ).json()
    recruit = client.post(
        f"/api/admin/journeys/{journey['id']}/recruits",
        headers=admin_headers(csrf),
        json={"name": "Concurrent Recruit"},
    ).json()
    first = {"items": [{"id": recruit["id"], "present": True, "arrival_time": None, "active": True, "base_version": recruit["version"]}]}
    assert client.put(f"/api/admin/journeys/{journey['id']}/attendance/recruits", headers=admin_headers(csrf), json=first).status_code == 200
    assert client.put(f"/api/admin/journeys/{journey['id']}/attendance/recruits", headers=admin_headers(csrf), json=first).status_code == 409
